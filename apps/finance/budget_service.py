import csv
import json
import re
from datetime import datetime, timezone
from io import TextIOWrapper
from pathlib import Path

from openpyxl import load_workbook

from .db import get_connection
from .service import FINANCE_IMPORTS_DIR, normalize_text


BUDGET_UNIT_PATTERN = re.compile(
    r"""
    (?P<fund>\d{4})
    -
    (?P<function>\d{4})
    -
    (?P<building>\d{3})
    -
    (?P<program>\d{3})
    -
    (?P<modifier>\d{2})
    -
    (?P<combined>\d+)
    (?:\s*-\s*(?P<title>.*))?
    """,
    re.VERBOSE,
)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_budget_unit(value: str | None) -> dict:
    value = normalize_text(value)

    result = {
        "ok": False,
        "fund_code": None,
        "function_code": None,
        "building_code": None,
        "program_code": None,
        "modifier_code": None,
        "combined_code": None,
        "title": None,
        "error": None,
    }

    if not value:
        result["error"] = "Budget Unit is blank."
        return result

    match = BUDGET_UNIT_PATTERN.search(value)

    if not match:
        result["error"] = "Budget Unit format was not recognized."
        return result

    result.update(
        {
            "ok": True,
            "fund_code": match.group("fund"),
            "function_code": match.group("function"),
            "building_code": match.group("building"),
            "program_code": match.group("program"),
            "modifier_code": match.group("modifier"),
            "combined_code": match.group("combined"),
            "title": normalize_text(match.group("title")),
            "error": None,
        }
    )

    return result


def find_budget_definition(
    *,
    combined_code: str | None,
    fiscal_year: str | None,
) -> dict | None:
    combined_code = normalize_text(combined_code)
    fiscal_year = normalize_text(fiscal_year)

    if not combined_code or not fiscal_year:
        return None

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_budget_definitions
            WHERE fiscal_year = ?
              AND combined_code = ?
              AND is_active = 1
            ORDER BY id DESC
            LIMIT 1
            """,
            (fiscal_year, combined_code),
        ).fetchone()

    return dict(row) if row else None


def get_active_budget_definition_set(fiscal_year: str | None = None) -> dict | None:
    fiscal_year = normalize_text(fiscal_year)

    with get_connection() as conn:
        if fiscal_year:
            row = conn.execute(
                """
                SELECT *
                FROM finance_budget_definition_sets
                WHERE fiscal_year = ?
                  AND is_active = 1
                ORDER BY uploaded_at DESC, id DESC
                LIMIT 1
                """,
                (fiscal_year,),
            ).fetchone()
        else:
            row = conn.execute(
                """
                SELECT *
                FROM finance_budget_definition_sets
                WHERE is_active = 1
                ORDER BY fiscal_year DESC, uploaded_at DESC, id DESC
                LIMIT 1
                """
            ).fetchone()

    return dict(row) if row else None


def enrich_budget_unit(
    *,
    budget_unit: str | None,
    fiscal_year: str | None = None,
) -> dict:
    parsed = parse_budget_unit(budget_unit)

    enriched = {
        **parsed,
        "budget_definition_id": None,
        "budget_definition_status": "parse_error",
        "definition": None,
    }

    if not parsed["ok"]:
        return enriched

    definition_set = get_active_budget_definition_set(fiscal_year)

    if not definition_set:
        enriched["budget_definition_status"] = "missing_definition_set"
        return enriched

    definition = find_budget_definition(
        combined_code=parsed["combined_code"],
        fiscal_year=definition_set["fiscal_year"],
    )

    if not definition:
        enriched["budget_definition_status"] = "missing_definition"
        return enriched

    enriched["budget_definition_id"] = definition["id"]
    enriched["budget_definition_status"] = "matched"
    enriched["definition"] = definition

    if definition.get("title"):
        enriched["title"] = definition["title"]

    return enriched


def _read_rows_from_import_file(stored_filename: str) -> list[dict]:
    file_path = FINANCE_IMPORTS_DIR / stored_filename

    if not file_path.exists():
        raise FileNotFoundError(f"Import file not found: {stored_filename}")

    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        with open(file_path, "rb") as raw_file:
            text_file = TextIOWrapper(raw_file, encoding="utf-8-sig", newline="")
            reader = csv.DictReader(text_file)
            return [
                {str(k).strip(): v for k, v in row.items() if k is not None}
                for row in reader
            ]

    if suffix == ".xlsx":
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        output = []

        for raw_row in rows[1:]:
            item = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = raw_row[index] if index < len(raw_row) else None
                item[header] = value
            output.append(item)

        return output

    raise ValueError("Unsupported file type. Only CSV and XLSX are supported.")


def _row_get(row: dict, *names: str):
    normalized = {
        str(key).strip().lower(): value
        for key, value in row.items()
        if key is not None
    }

    for name in names:
        key = name.strip().lower()
        if key in normalized:
            return normalized[key]

    return None


def import_budget_definitions(
    *,
    stored_filename: str,
    original_filename: str,
    fiscal_year: str,
    uploaded_by_user_id: int | None = None,
    notes: str | None = None,
) -> dict:
    fiscal_year = normalize_text(fiscal_year)
    if not fiscal_year:
        raise ValueError("Fiscal year is required.")

    rows = _read_rows_from_import_file(stored_filename)
    now = utc_now_iso()

    imported_count = 0
    skipped_count = 0

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_budget_definition_sets
            SET is_active = 0,
                updated_at = ?
            WHERE fiscal_year = ?
            """,
            (now, fiscal_year),
        )

        cursor = conn.execute(
            """
            INSERT INTO finance_budget_definition_sets (
                fiscal_year,
                source_filename,
                uploaded_by_user_id,
                uploaded_at,
                is_active,
                notes,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, 1, ?, ?, ?)
            """,
            (
                fiscal_year,
                normalize_text(original_filename),
                uploaded_by_user_id,
                now,
                normalize_text(notes),
                now,
                now,
            ),
        )

        definition_set_id = cursor.lastrowid

        for row in rows:
            level_number = normalize_text(str(_row_get(row, "Level", "LEVEL", "level") or ""))
            code = normalize_text(str(_row_get(row, "Code", "CODE", "code") or ""))
            title = normalize_text(str(_row_get(row, "Title", "TITLE", "Description", "DESCRIPTION") or ""))

            fund_code = normalize_text(str(_row_get(row, "Fund", "FUND", "Level 1") or ""))
            function_code = normalize_text(str(_row_get(row, "Function", "FUNCTION", "Level 2") or ""))
            building_code = normalize_text(str(_row_get(row, "Building", "BUILDING", "Level 3") or ""))
            program_code = normalize_text(str(_row_get(row, "Program", "PROGRAM", "Level 4") or ""))
            modifier_code = normalize_text(str(_row_get(row, "Modifier", "MODIFIER", "Level 5") or ""))
            combined_code = normalize_text(str(_row_get(row, "Combined", "COMBINED", "Budget Unit", "BUDGET UNIT", "Level 6") or ""))

            # Support eFinance rows where Level 6 is the combined account code.
            if level_number == "6" and code and not combined_code:
                combined_code = code

            # Support rows where one column contains the full dashed budget unit.
            possible_budget_unit = normalize_text(
                str(_row_get(row, "Budget Unit", "BUDGET UNIT", "budget_unit") or "")
            )
            parsed = parse_budget_unit(possible_budget_unit)

            if parsed["ok"]:
                fund_code = fund_code or parsed["fund_code"]
                function_code = function_code or parsed["function_code"]
                building_code = building_code or parsed["building_code"]
                program_code = program_code or parsed["program_code"]
                modifier_code = modifier_code or parsed["modifier_code"]
                combined_code = combined_code or parsed["combined_code"]
                title = title or parsed["title"]

            if not combined_code and not code:
                skipped_count += 1
                continue

            try:
                clean_level_number = int(level_number) if level_number else None
            except ValueError:
                clean_level_number = None

            conn.execute(
                """
                INSERT INTO finance_budget_definitions (
                    definition_set_id,
                    fiscal_year,
                    level_number,
                    code,
                    fund_code,
                    function_code,
                    building_code,
                    program_code,
                    modifier_code,
                    combined_code,
                    title,
                    description,
                    raw_json,
                    is_active,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                """,
                (
                    definition_set_id,
                    fiscal_year,
                    clean_level_number,
                    code,
                    fund_code,
                    function_code,
                    building_code,
                    program_code,
                    modifier_code,
                    combined_code,
                    title,
                    title,
                    json.dumps(row, default=str),
                    now,
                    now,
                ),
            )

            imported_count += 1

        conn.commit()

    return {
        "definition_set_id": definition_set_id,
        "fiscal_year": fiscal_year,
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "total_rows": len(rows),
    }


def get_budget_definition_summary() -> dict:
    with get_connection() as conn:
        active_sets = conn.execute(
            """
            SELECT
                s.*,
                COUNT(d.id) AS definition_count
            FROM finance_budget_definition_sets s
            LEFT JOIN finance_budget_definitions d
                ON d.definition_set_id = s.id
            WHERE s.is_active = 1
            GROUP BY s.id
            ORDER BY s.fiscal_year DESC, s.uploaded_at DESC
            """
        ).fetchall()

        missing_transactions = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM finance_transactions
            WHERE budget_definition_status IN ('missing_definition', 'missing_definition_set', 'parse_error')
            """
        ).fetchone()["count"]

    return {
        "active_sets": [dict(row) for row in active_sets],
        "missing_transactions": missing_transactions,
    }