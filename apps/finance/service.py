from datetime import datetime, timezone, date
from apps.finance.notification_service import send_finance_test_email
from modules.core.settings.settings_service import get_setting, get_bool_setting
from modules.core.identity.identity_db import get_connection as get_identity_connection
from .db import get_connection
import os
import secrets
import json
from pathlib import Path
import csv
from io import TextIOWrapper
from openpyxl import load_workbook
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher

BASE_DIR = Path(__file__).resolve().parents[2]
FINANCE_UPLOADS_DIR = BASE_DIR / "instance" / "finance" / "uploads"
FINANCE_IMPORTS_DIR = BASE_DIR / "instance" / "finance" / "imports"

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def ensure_finance_imports_dir():
    FINANCE_IMPORTS_DIR.mkdir(parents=True, exist_ok=True)

def normalize_text(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return value or None


def ensure_finance_department(department_name: str) -> dict:
    department_name = normalize_text(department_name)
    if not department_name:
        raise ValueError("department_name is required")

    now = utc_now_iso()

    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT *
            FROM finance_departments
            WHERE department_name = ?
            """,
            (department_name,),
        ).fetchone()

        if existing:
            return dict(existing)

        conn.execute(
            """
            INSERT INTO finance_departments (
                department_name,
                is_enabled,
                created_at,
                updated_at
            )
            VALUES (?, 1, ?, ?)
            """,
            (department_name, now, now),
        )
        conn.commit()

        row = conn.execute(
            """
            SELECT *
            FROM finance_departments
            WHERE department_name = ?
            """,
            (department_name,),
        ).fetchone()

    return dict(row)


def list_active_departments_from_users() -> list[str]:
    with get_identity_connection() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT TRIM(department) AS department_name
            FROM users
            WHERE is_active = 1
              AND department IS NOT NULL
              AND TRIM(department) <> ''
            ORDER BY department_name COLLATE NOCASE
            """
        ).fetchall()

    return [row["department_name"] for row in rows if row["department_name"]]


def sync_departments_from_users() -> list[dict]:
    for department_name in list_active_departments_from_users():
        ensure_finance_department(department_name)

    return list_enabled_departments()


def list_enabled_departments() -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_departments
            WHERE is_enabled = 1
            ORDER BY department_name COLLATE NOCASE
            """
        ).fetchall()

    return [dict(row) for row in rows]


def get_department_record(department_name: str) -> dict | None:
    department_name = normalize_text(department_name)
    if not department_name:
        return None

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_departments
            WHERE department_name = ?
            """
            ,
            (department_name,),
        ).fetchone()

    return dict(row) if row else None


def list_vendors() -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_vendors
            WHERE is_active = 1
            ORDER BY vendor_name COLLATE NOCASE
            """
        ).fetchall()

    return [dict(row) for row in rows]

def list_records_for_vendor(vendor_id: int) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                r.*,
                c.category_name
            FROM finance_records r
            LEFT JOIN finance_categories c ON c.id = r.category_id
            WHERE r.vendor_id = ?
            ORDER BY
                CASE WHEN r.renewal_date IS NULL THEN 1 ELSE 0 END,
                r.renewal_date ASC,
                r.title COLLATE NOCASE ASC
            """,
            (vendor_id,),
        ).fetchall()

    return [dict(row) for row in rows]


def list_active_vendors(
    q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    q = normalize_text(q)

    page = max(int(page or 1), 1)
    per_page = max(min(int(per_page or 100), 250), 25)
    offset = (page - 1) * per_page

    where = ["status IS NULL OR status NOT IN ('archived', 'deleted')"]
    params = []

    if q:
        where.append(
            """
            (
                LOWER(COALESCE(vendor_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(vendor_code, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(website, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(billing_email, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(support_email, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(sales_contact_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(notes, '')) LIKE LOWER(?)
            )
            """
        )
        term = f"%{q}%"
        params.extend([term, term, term, term, term, term, term])

    where_sql = " AND ".join(f"({item})" for item in where)

    with get_connection() as conn:
        total = conn.execute(
            f"""
            SELECT COUNT(*) AS count
            FROM finance_vendors
            WHERE {where_sql}
            """,
            params,
        ).fetchone()["count"]

        rows = conn.execute(
            f"""
            SELECT *
            FROM finance_vendors
            WHERE {where_sql}
            ORDER BY vendor_name COLLATE NOCASE ASC
            LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()

    total_pages = max((total + per_page - 1) // per_page, 1)

    return {
        "rows": [dict(row) for row in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": page - 1,
        "next_page": page + 1,
    }


def list_archived_vendors(
    q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    q = normalize_text(q)

    page = max(int(page or 1), 1)
    per_page = max(min(int(per_page or 100), 250), 25)
    offset = (page - 1) * per_page

    where = ["status = 'archived'"]
    params = []

    if q:
        where.append("""
            (
                LOWER(COALESCE(vendor_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(vendor_code, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(website, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(billing_email, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(support_email, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(sales_contact_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(notes, '')) LIKE LOWER(?)
            )
        """)
        term = f"%{q}%"
        params.extend([term, term, term, term, term, term, term])

    where_sql = " AND ".join(f"({item})" for item in where)

    with get_connection() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) AS count FROM finance_vendors WHERE {where_sql}",
            params,
        ).fetchone()["count"]

        rows = conn.execute(
            f"""
            SELECT *
            FROM finance_vendors
            WHERE {where_sql}
            ORDER BY vendor_name COLLATE NOCASE ASC
            LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()

    return {
        "rows": [dict(row) for row in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max((total + per_page - 1) // per_page, 1),
        "has_prev": page > 1,
        "has_next": page < max((total + per_page - 1) // per_page, 1),
        "prev_page": page - 1,
        "next_page": page + 1,
    }


def list_deleted_vendors(
    q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    q = normalize_text(q)

    page = max(int(page or 1), 1)
    per_page = max(min(int(per_page or 100), 250), 25)
    offset = (page - 1) * per_page

    where = ["status = 'deleted'"]
    params = []

    if q:
        where.append("""
            (
                LOWER(COALESCE(vendor_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(vendor_code, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(website, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(billing_email, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(support_email, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(sales_contact_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(notes, '')) LIKE LOWER(?)
            )
        """)
        term = f"%{q}%"
        params.extend([term, term, term, term, term, term, term])

    where_sql = " AND ".join(f"({item})" for item in where)

    with get_connection() as conn:
        total = conn.execute(
            f"""
            SELECT COUNT(*) AS count
            FROM finance_vendors
            WHERE {where_sql}
            """,
            params,
        ).fetchone()["count"]

        rows = conn.execute(
            f"""
            SELECT *
            FROM finance_vendors
            WHERE {where_sql}
            ORDER BY deleted_at DESC, vendor_name COLLATE NOCASE ASC
            LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()

    total_pages = max((total + per_page - 1) // per_page, 1)

    return {
        "rows": [dict(row) for row in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": page - 1,
        "next_page": page + 1,
    }


def archive_vendor(vendor_id: int) -> bool:
    vendor = get_vendor_by_id(vendor_id)
    if not vendor:
        return False

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_vendors
            SET
                status = 'archived',
                is_active = 0,
                updated_at = ?
            WHERE id = ?
            """,
            (now, vendor_id),
        )
        conn.commit()

    return True


def delete_vendor(vendor_id: int) -> bool:
    vendor = get_vendor_by_id(vendor_id)
    if not vendor:
        return False

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_vendors
            SET
                status = 'deleted',
                deleted_at = ?,
                is_active = 0,
                updated_at = ?
            WHERE id = ?
            """,
            (now, now, vendor_id),
        )
        conn.commit()

    return True


def restore_vendor(vendor_id: int) -> bool:
    vendor = get_vendor_by_id(vendor_id)
    if not vendor:
        return False

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_vendors
            SET
                status = 'active',
                deleted_at = NULL,
                is_active = 1,
                updated_at = ?
            WHERE id = ?
            """,
            (now, vendor_id),
        )
        conn.commit()

    return True


def list_categories() -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_categories
            WHERE is_active = 1
            ORDER BY sort_order, category_name COLLATE NOCASE
            """
        ).fetchall()

    return [dict(row) for row in rows]


def get_finance_dashboard_summary(department_name: str) -> dict:
    department_name = normalize_text(department_name)
    if not department_name:
        return {
            "total_records": 0,
            "renewals_due_30": 0,
            "active_vendors": 0,
        }

    with get_connection() as conn:
        total_records = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM finance_records
            WHERE department_name = ?
              AND status NOT IN ('archived', 'deleted')
            """,
            (department_name,),
        ).fetchone()["count"]

        renewals_due_30 = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM finance_records
            WHERE department_name = ?
              AND renewal_date IS NOT NULL
              AND status IN ('active', 'pending_renewal')
              AND DATE(renewal_date) <= DATE('now', '+30 day')
            """,
            (department_name,),
        ).fetchone()["count"]

        active_vendors = conn.execute(
            """
            SELECT COUNT(DISTINCT vendor_id) AS count
            FROM finance_records
            WHERE department_name = ?
              AND vendor_id IS NOT NULL
              AND status NOT IN ('archived', 'deleted')
            """,
            (department_name,),
        ).fetchone()["count"]

    return {
        "total_records": total_records,
        "renewals_due_30": renewals_due_30,
        "active_vendors": active_vendors,
    }

def create_record(
    *,
    record_type: str,
    title: str,
    department_name: str,
    vendor_id: int | None = None,
    category_id: int | None = None,
    account_code: str | None = None,
    po_number: str | None = None,
    purchase_date: str | None = None,
    service_start_date: str | None = None,
    use_purchase_date_as_start: bool = True,
    term_length: int | None = None,
    term_unit: str | None = None,
    expiration_date: str | None = None,
    renewal_date: str | None = None,
    notify_days_before: int = 30,
    notification_recipients: str | None = None,
    status: str = "active",
    cost: str | None = None,
    notes: str | None = None,
    created_by_user_id: int | None = None,
) -> int:
    title = normalize_text(title)
    department_name = normalize_text(department_name)
    record_type = normalize_text(record_type) or "renewal"
    status = (normalize_text(status) or "active").lower()

    if not title:
        raise ValueError("title is required")
    if not department_name:
        raise ValueError("department_name is required")

    now = utc_now_iso()

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO finance_records (
                record_type,
                title,
                vendor_id,
                category_id,
                department_name,
                account_code,
                po_number,
                purchase_date,
                service_start_date,
                use_purchase_date_as_start,
                term_length,
                term_unit,
                expiration_date,
                renewal_date,
                notify_days_before,
                notification_recipients,
                status,
                cost,
                notes,
                created_by_user_id,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record_type,
                title,
                vendor_id,
                category_id,
                department_name,
                normalize_text(account_code),
                normalize_text(po_number),
                normalize_text(purchase_date),
                normalize_text(service_start_date),
                1 if use_purchase_date_as_start else 0,
                term_length,
                normalize_text(term_unit),
                normalize_text(expiration_date),
                normalize_text(renewal_date),
                notify_days_before or 30,
                normalize_text(notification_recipients),
                status,
                normalize_text(cost),
                normalize_text(notes),
                created_by_user_id,
                now,
                now,
            ),
        )
        record_id = cursor.lastrowid

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'created', ?, ?, ?)
            """,
            (
                record_id,
                f"Record created: {title}",
                created_by_user_id,
                now,
            ),
        )

        conn.commit()
        return record_id
    

def list_records_for_department(department_name: str) -> list[dict]:
    department_name = normalize_text(department_name)
    if not department_name:
        return []

    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                r.*,
                v.vendor_name,
                c.category_name
            FROM finance_records r
            LEFT JOIN finance_vendors v ON v.id = r.vendor_id
            LEFT JOIN finance_categories c ON c.id = r.category_id
            WHERE r.department_name = ?
            ORDER BY
                CASE WHEN r.renewal_date IS NULL THEN 1 ELSE 0 END,
                r.renewal_date ASC,
                r.title COLLATE NOCASE ASC
            """,
            (department_name,),
        ).fetchall()

    return [dict(row) for row in rows]


def get_record_by_id(record_id: int) -> dict | None:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT
                r.*,
                v.vendor_name,
                c.category_name
            FROM finance_records r
            LEFT JOIN finance_vendors v ON v.id = r.vendor_id
            LEFT JOIN finance_categories c ON c.id = r.category_id
            WHERE r.id = ?
            """,
            (record_id,),
        ).fetchone()

    return dict(row) if row else None

def _parse_iso_date(value: str | None) -> date | None:
    value = normalize_text(value)
    if not value:
        return None

    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def should_send_renewal_notification_now(
    *,
    renewal_date: str | None,
    notify_days_before: int | None,
    today: date | None = None,
    overdue_grace_days: int = 30,
) -> bool:
    renewal = _parse_iso_date(renewal_date)
    if not renewal:
        return False

    if today is None:
        today = date.today()

    try:
        days_before = int(notify_days_before or 0)
    except (TypeError, ValueError):
        days_before = 0

    if days_before < 0:
        days_before = 0

    delta_days = (renewal - today).days
    return (-overdue_grace_days) <= delta_days <= days_before


def build_finance_notification_preview_lines(record: dict) -> list[dict]:
    lines = []

    field_map = [
        ("title", "Title", False),
        ("department_name", "Department", False),
        ("vendor_name", "Vendor", False),
        ("category_name", "Category", False),
        ("renewal_date", "Renewal Date", False),
        ("expiration_date", "Expiration Date", False),
        ("cost", "Cost", False),
        ("po_number", "PO Number", False),
        ("account_code", "Account Code", False),
        ("notes", "Notes", False),
        ("record_url", "Record Link", True),
    ]

    for key, label, is_link in field_map:
        value = record.get(key)
        if value is None or str(value).strip() == "":
            continue

        lines.append(
            {
                "label": label,
                "value": value,
                "is_link": is_link,
            }
        )

    return lines


def build_finance_notification_context(record: dict) -> dict:
    record_id = record["id"]
    vendor_name = record.get("vendor_name") or ""
    category_name = record.get("category_name") or ""

    if not vendor_name and record.get("vendor_id"):
        vendor = get_vendor_by_id(record["vendor_id"])
        vendor_name = vendor["vendor_name"] if vendor else ""

    if not category_name and record.get("category_id"):
        categories = list_categories()
        category_map = {item["id"]: item["category_name"] for item in categories}
        category_name = category_map.get(record["category_id"], "")

    renewal = _parse_iso_date(record.get("renewal_date"))
    today = date.today()
    days_until_renewal = ""
    if renewal:
        days_until_renewal = str((renewal - today).days)

    record_url = f"/records/{record_id}"
    public_base_url = (get_setting("general.public_base_url", "") or "").strip().rstrip("/")
    if public_base_url:
        record_url = f"{public_base_url}/finance/records/{record_id}"

    context = {
        "title": record.get("title") or "",
        "department_name": record.get("department_name") or "",
        "vendor_name": vendor_name,
        "category_name": category_name,
        "renewal_date": record.get("renewal_date") or "",
        "expiration_date": record.get("expiration_date") or "",
        "cost": record.get("cost") or "",
        "po_number": record.get("po_number") or "",
        "account_code": record.get("account_code") or "",
        "notification_recipients": record.get("notification_recipients") or "",
        "notes": record.get("notes") or "",
        "record_url": record_url,
        "days_until_renewal": days_until_renewal,
    }

    return context


def maybe_send_renewal_notification_for_record(
    record_id: int,
    *,
    changed_by_user_id: int | None = None,
) -> tuple[bool, str]:
    record = get_record_by_id(record_id)
    if not record:
        return False, "Record not found."

    notifications_enabled = get_bool_setting("finance.notifications.enabled", False)
    if not notifications_enabled:
        return False, "Finance notifications are disabled."

    record_status = (record.get("status") or "").strip().lower()
    if record_status not in {"active", "pending_renewal"}:
        return False, "Record status is not eligible for renewal notifications."

    renewal_date = record.get("renewal_date")
    if not should_send_renewal_notification_now(
        renewal_date=renewal_date,
        notify_days_before=record.get("notify_days_before"),
    ):
        return False, "Record is not within its notification window."

    if has_renewal_notification_been_sent(
        record_id=record_id,
        renewal_date=renewal_date,
    ):
        return False, "Renewal notification has already been sent for this renewal date."

    sender_email = (get_setting("finance.notifications.sender_email", "") or "").strip()
    default_recipients = (get_setting("finance.notifications.default_recipients", "") or "").strip()
    use_record_recipients_first = get_bool_setting(
        "finance.notifications.use_record_recipients_first",
        True,
    )
    fallback_to_default_recipients = get_bool_setting(
        "finance.notifications.fallback_to_default_recipients",
        True,
    )

    record_recipients = (record.get("notification_recipients") or "").strip()

    recipient_email = ""
    if use_record_recipients_first and record_recipients:
        recipient_email = record_recipients
    elif default_recipients:
        recipient_email = default_recipients

    if not recipient_email and fallback_to_default_recipients and default_recipients:
        recipient_email = default_recipients

    if not sender_email:
        return False, "Finance sender email is not configured."

    if not recipient_email:
        return False, "No notification recipient is configured for this record."

    template_header = (get_setting("finance.notifications.template_header", "") or "").strip()
    template_intro = (get_setting("finance.notifications.template_intro", "") or "").strip()
    template_subject = (get_setting("finance.notifications.template_subject", "") or "").strip()
    template_footer = (get_setting("finance.notifications.template_footer", "") or "").strip()
    subject_prefix = (get_setting("finance.notifications.subject_prefix", "") or "").strip()

    preview_context = build_finance_notification_context(record)
    preview_lines = build_finance_notification_preview_lines(preview_context)

    send_finance_test_email(
        sender_email=sender_email,
        recipient_email=recipient_email,
        subject_template=template_subject,
        subject_prefix=subject_prefix,
        template_header=template_header,
        template_intro=template_intro,
        template_footer=template_footer,
        preview_context=preview_context,
        preview_lines=preview_lines,
    )

    log_renewal_notification_sent(
        record_id=record_id,
        renewal_date=renewal_date,
        recipient_email=recipient_email,
        changed_by_user_id=changed_by_user_id,
    )

    return True, f"Notification sent to {recipient_email}."

def has_renewal_notification_been_sent(
    *,
    record_id: int,
    renewal_date: str | None,
) -> bool:
    renewal_date = normalize_text(renewal_date)
    if not renewal_date:
        return False

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT id
            FROM finance_record_history
            WHERE finance_record_id = ?
              AND event_type = 'renewal_notification_sent'
              AND summary LIKE ?
            LIMIT 1
            """,
            (record_id, f"Renewal notification sent for renewal date {renewal_date}%"),
        ).fetchone()

    return row is not None


def log_renewal_notification_sent(
    *,
    record_id: int,
    renewal_date: str | None,
    recipient_email: str,
    changed_by_user_id: int | None = None,
):
    renewal_date = normalize_text(renewal_date)
    recipient_email = normalize_text(recipient_email)

    if not renewal_date:
        return

    now = utc_now_iso()
    summary = f"Renewal notification sent for renewal date {renewal_date}"
    if recipient_email:
        summary = f"{summary} to {recipient_email}"

    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'renewal_notification_sent', ?, ?, ?)
            """,
            (
                record_id,
                summary,
                changed_by_user_id,
                now,
            ),
        )
        conn.commit()

def create_vendor(
    vendor_name: str,
    vendor_code: str | None = None,
    website: str | None = None,
    main_phone: str | None = None,
    billing_email: str | None = None,
    support_email: str | None = None,
    sales_contact_name: str | None = None,
    sales_contact_email: str | None = None,
    notes: str | None = None,
) -> int:
    vendor_name = normalize_text(vendor_name)
    if not vendor_name:
        raise ValueError("vendor_name is required")

    now = utc_now_iso()

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO finance_vendors (
                vendor_name,
                vendor_code,
                website,
                main_phone,
                billing_email,
                support_email,
                sales_contact_name,
                sales_contact_email,
                is_active,
                status,
                notes,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, 'active', ?, ?, ?)
            """,
            (
                vendor_name,
                normalize_text(vendor_code),
                normalize_text(website),
                normalize_text(main_phone),
                normalize_text(billing_email),
                normalize_text(support_email),
                normalize_text(sales_contact_name),
                normalize_text(sales_contact_email),
                normalize_text(notes),
                now,
                now,
            ),
        )
        conn.commit()
        return cursor.lastrowid


def update_vendor(
    vendor_id: int,
    vendor_name: str,
    vendor_code: str | None = None,
    website: str | None = None,
    main_phone: str | None = None,
    billing_email: str | None = None,
    support_email: str | None = None,
    sales_contact_name: str | None = None,
    sales_contact_email: str | None = None,
    status: str = "active",
    notes: str | None = None,
):
    vendor_name = normalize_text(vendor_name)
    if not vendor_name:
        raise ValueError("vendor_name is required")

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_vendors
            SET
                vendor_name = ?,
                vendor_code = ?,
                website = ?,
                main_phone = ?,
                billing_email = ?,
                support_email = ?,
                sales_contact_name = ?,
                sales_contact_email = ?,
                is_active = ?,
                status = ?,
                notes = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                vendor_name,
                normalize_text(vendor_code),
                normalize_text(website),
                normalize_text(main_phone),
                normalize_text(billing_email),
                normalize_text(support_email),
                normalize_text(sales_contact_name),
                normalize_text(sales_contact_email),
                1 if status == "active" else 0,
                normalize_text(status) or "active",
                normalize_text(notes),
                now,
                vendor_id,
            ),
        )
        conn.commit()


def get_vendor_by_id(vendor_id: int) -> dict | None:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_vendors
            WHERE id = ?
            """,
            (vendor_id,),
        ).fetchone()

    return dict(row) if row else None


def list_vendors_all() -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_vendors
            ORDER BY vendor_name COLLATE NOCASE
            """
        ).fetchall()

    return [dict(row) for row in rows]

def normalize_vendor_name(value: str | None) -> str | None:
    value = normalize_text(value)
    if not value:
        return None

    v = value.strip()

    # preserve short acronyms (SHI, CDW, B&H)
    if v.isupper() and len(v) <= 5:
        return v

    # preserve mixed acronyms like CDW-G
    if "-" in v and v.upper() == v:
        return v

    # normal title case
    return " ".join(word.capitalize() for word in v.split())

def update_record(
    *,
    record_id: int,
    record_type: str,
    title: str,
    department_name: str,
    vendor_id: int | None = None,
    category_id: int | None = None,
    account_code: str | None = None,
    po_number: str | None = None,
    purchase_date: str | None = None,
    service_start_date: str | None = None,
    use_purchase_date_as_start: bool = True,
    term_length: int | None = None,
    term_unit: str | None = None,
    expiration_date: str | None = None,
    renewal_date: str | None = None,
    notify_days_before: int = 30,
    notification_recipients: str | None = None,
    status: str = "active",
    cost: str | None = None,
    notes: str | None = None,
    changed_by_user_id: int | None = None,
):
    title = normalize_text(title)
    department_name = normalize_text(department_name)
    record_type = normalize_text(record_type) or "renewal"
    status = (normalize_text(status) or "active").lower()

    if not title:
        raise ValueError("title is required")
    if not department_name:
        raise ValueError("department_name is required")

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_records
            SET
                record_type = ?,
                title = ?,
                vendor_id = ?,
                category_id = ?,
                department_name = ?,
                account_code = ?,
                po_number = ?,
                purchase_date = ?,
                service_start_date = ?,
                use_purchase_date_as_start = ?,
                term_length = ?,
                term_unit = ?,
                expiration_date = ?,
                renewal_date = ?,
                notify_days_before = ?,
                notification_recipients = ?,
                status = ?,
                cost = ?,
                notes = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                record_type,
                title,
                vendor_id,
                category_id,
                department_name,
                normalize_text(account_code),
                normalize_text(po_number),
                normalize_text(purchase_date),
                normalize_text(service_start_date),
                1 if use_purchase_date_as_start else 0,
                term_length,
                normalize_text(term_unit),
                normalize_text(expiration_date),
                normalize_text(renewal_date),
                notify_days_before or 30,
                normalize_text(notification_recipients),
                status,
                normalize_text(cost),
                normalize_text(notes),
                now,
                record_id,
            ),
        )

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'updated', ?, ?, ?)
            """,
            (
                record_id,
                f"Record updated: {title}",
                changed_by_user_id,
                now,
            ),
        )

        conn.commit()

def ensure_finance_uploads_dir():
    FINANCE_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def list_attachments_for_record(record_id: int) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_attachments
            WHERE finance_record_id = ?
            ORDER BY uploaded_at DESC, id DESC
            """,
            (record_id,),
        ).fetchall()

    return [dict(row) for row in rows]


def get_attachment_by_id(attachment_id: int) -> dict | None:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_attachments
            WHERE id = ?
            """,
            (attachment_id,),
        ).fetchone()

    return dict(row) if row else None


ALLOWED_ATTACHMENT_EXTENSIONS = {
    ".pdf",
    ".csv",
    ".xml",
    ".txt",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
}

def save_attachment(
    *,
    finance_record_id: int,
    original_filename: str,
    file_bytes: bytes,
    mime_type: str,
    document_type: str = "other",
    uploaded_by_user_id: int | None = None,
) -> int:
    if not original_filename:
        raise ValueError("original_filename is required")
    if not file_bytes:
        raise ValueError("file_bytes is required")

    suffix = Path(original_filename).suffix.lower()
    if suffix not in ALLOWED_ATTACHMENT_EXTENSIONS:
        raise ValueError(
            "Unsupported attachment type. Allowed types: "
            "PDF, CSV, XML, TXT, DOC, DOCX, XLS, XLSX, JPG, JPEG, PNG, WEBP."
        )

    ensure_finance_uploads_dir()

    stored_name = f"{secrets.token_hex(16)}{suffix}"
    stored_path = FINANCE_UPLOADS_DIR / stored_name

    with open(stored_path, "wb") as f:
        f.write(file_bytes)

    now = utc_now_iso()
    file_size = len(file_bytes)

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO finance_attachments (
                finance_record_id,
                file_name,
                stored_name,
                mime_type,
                file_size,
                document_type,
                uploaded_by_user_id,
                uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                finance_record_id,
                original_filename,
                stored_name,
                mime_type or "application/octet-stream",
                file_size,
                normalize_text(document_type) or "other",
                uploaded_by_user_id,
                now,
            ),
        )

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'attachment_added', ?, ?, ?)
            """,
            (
                finance_record_id,
                f"Attachment added: {original_filename}",
                uploaded_by_user_id,
                now,
            ),
        )

        conn.commit()
        return cursor.lastrowid
    
def delete_attachment(attachment_id: int, deleted_by_user_id: int | None = None) -> bool:
    attachment = get_attachment_by_id(attachment_id)
    if not attachment:
        return False

    file_path = FINANCE_UPLOADS_DIR / attachment["stored_name"]
    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            DELETE FROM finance_attachments
            WHERE id = ?
            """,
            (attachment_id,),
        )

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'attachment_deleted', ?, ?, ?)
            """,
            (
                attachment["finance_record_id"],
                f"Attachment deleted: {attachment['file_name']}",
                deleted_by_user_id,
                now,
            ),
        )

        conn.commit()

    if file_path.exists():
        try:
            file_path.unlink()
        except OSError:
            pass

    return True

def list_history_for_record(record_id: int) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_record_history
            WHERE finance_record_id = ?
            ORDER BY changed_at DESC, id DESC
            """,
            (record_id,),
        ).fetchall()

    return [dict(row) for row in rows]


def archive_record(record_id: int, changed_by_user_id: int | None = None) -> bool:
    record = get_record_by_id(record_id)
    if not record:
        return False

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_records
            SET
                status = 'archived',
                updated_at = ?
            WHERE id = ?
            """,
            (now, record_id),
        )

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'archived', ?, ?, ?)
            """,
            (
                record_id,
                f"Record archived: {record['title']}",
                changed_by_user_id,
                now,
            ),
        )

        conn.commit()

    return True


def delete_record(record_id: int, deleted_by_user_id: int | None = None) -> bool:
    record = get_record_by_id(record_id)
    if not record:
        return False

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_records
            SET
                status = 'deleted',
                deleted_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (now, now, record_id),
        )

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'deleted', ?, ?, ?)
            """,
            (
                record_id,
                f"Record moved to deleted: {record['title']}",
                deleted_by_user_id,
                now,
            ),
        )

        conn.commit()

    return True

def list_records_for_department_page(
    *,
    department_name: str,
    status: str,
    q: str | None = None,
    category_id: int | None = None,
    vendor_q: str | None = None,
    record_type: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    department_name = normalize_text(department_name)
    q = normalize_text(q)
    vendor_q = normalize_text(vendor_q)
    record_type = normalize_text(record_type)

    page = max(int(page or 1), 1)
    per_page = max(min(int(per_page or 100), 250), 25)
    offset = (page - 1) * per_page

    if not department_name:
        return {
            "rows": [],
            "total": 0,
            "page": page,
            "per_page": per_page,
            "total_pages": 1,
            "has_prev": False,
            "has_next": False,
            "prev_page": page - 1,
            "next_page": page + 1,
        }

    where = ["r.department_name = ?"]
    params = [department_name]

    if status == "active":
        where.append("r.status NOT IN ('archived', 'deleted')")
    elif status == "archived":
        where.append("r.status = 'archived'")
    elif status == "deleted":
        where.append("r.status = 'deleted'")
    else:
        raise ValueError("Invalid record status.")

    if q:
        where.append("""
            (
                LOWER(COALESCE(r.title, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(v.vendor_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(c.category_name, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(r.account_code, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(r.po_number, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(r.notes, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(r.record_type, '')) LIKE LOWER(?)
                OR LOWER(COALESCE(r.cost, '')) LIKE LOWER(?)
            )
        """)
        term = f"%{q}%"
        params.extend([term, term, term, term, term, term, term, term])

    if category_id:
        where.append("r.category_id = ?")
        params.append(category_id)

    if vendor_q:
        where.append("LOWER(COALESCE(v.vendor_name, '')) LIKE LOWER(?)")
        params.append(f"%{vendor_q}%")

    if record_type:
        where.append("LOWER(COALESCE(r.record_type, '')) = LOWER(?)")
        params.append(record_type)

    where_sql = " AND ".join(f"({item})" for item in where)

    if status == "deleted":
        order_sql = "r.deleted_at DESC, r.title COLLATE NOCASE ASC"
    elif status == "archived":
        order_sql = "r.updated_at DESC, r.title COLLATE NOCASE ASC"
    else:
        order_sql = """
            CASE WHEN r.renewal_date IS NULL THEN 1 ELSE 0 END,
            r.renewal_date ASC,
            r.title COLLATE NOCASE ASC
        """

    with get_connection() as conn:
        total = conn.execute(
            f"""
            SELECT COUNT(*) AS count
            FROM finance_records r
            LEFT JOIN finance_vendors v ON v.id = r.vendor_id
            LEFT JOIN finance_categories c ON c.id = r.category_id
            WHERE {where_sql}
            """,
            params,
        ).fetchone()["count"]

        rows = conn.execute(
            f"""
            SELECT
                r.*,
                v.vendor_name,
                c.category_name
            FROM finance_records r
            LEFT JOIN finance_vendors v ON v.id = r.vendor_id
            LEFT JOIN finance_categories c ON c.id = r.category_id
            WHERE {where_sql}
            ORDER BY {order_sql}
            LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()

    total_pages = max((total + per_page - 1) // per_page, 1)

    return {
        "rows": [dict(row) for row in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": page - 1,
        "next_page": page + 1,
    }


def list_active_records_for_department(
    department_name: str,
    q: str | None = None,
    category_id: int | None = None,
    vendor_q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    return list_records_for_department_page(
        department_name=department_name,
        status="active",
        q=q,
        category_id=category_id,
        vendor_q=vendor_q,
        page=page,
        per_page=per_page,
    )

def list_renewal_records_for_department(
    department_name: str,
    q: str | None = None,
    category_id: int | None = None,
    vendor_q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    return list_records_for_department_page(
        department_name=department_name,
        status="active",
        q=q,
        category_id=category_id,
        vendor_q=vendor_q,
        record_type="renewal",
        page=page,
        per_page=per_page,
    )


def list_archived_records_for_department(
    department_name: str,
    q: str | None = None,
    category_id: int | None = None,
    vendor_q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    return list_records_for_department_page(
        department_name=department_name,
        status="archived",
        q=q,
        category_id=category_id,
        vendor_q=vendor_q,
        page=page,
        per_page=per_page,
    )


def list_deleted_records_for_department(
    department_name: str,
    q: str | None = None,
    category_id: int | None = None,
    vendor_q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    return list_records_for_department_page(
        department_name=department_name,
        status="deleted",
        q=q,
        category_id=category_id,
        vendor_q=vendor_q,
        page=page,
        per_page=per_page,
    )

def restore_archived_record(record_id: int, changed_by_user_id: int | None = None) -> bool:
    record = get_record_by_id(record_id)
    if not record:
        return False

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_records
            SET
                status = 'active',
                updated_at = ?
            WHERE id = ?
            """,
            (now, record_id),
        )

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'restored', ?, ?, ?)
            """,
            (
                record_id,
                f"Archived record restored: {record['title']}",
                changed_by_user_id,
                now,
            ),
        )

        conn.commit()

    return True

def restore_deleted_record(record_id: int, changed_by_user_id: int | None = None) -> bool:
    record = get_record_by_id(record_id)
    if not record:
        return False

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_records
            SET
                status = 'active',
                deleted_at = NULL,
                updated_at = ?
            WHERE id = ?
            """,
            (now, record_id),
        )

        conn.execute(
            """
            INSERT INTO finance_record_history (
                finance_record_id,
                event_type,
                summary,
                changed_by_user_id,
                changed_at
            )
            VALUES (?, 'restored', ?, ?, ?)
            """,
            (
                record_id,
                f"Deleted record restored: {record['title']}",
                changed_by_user_id,
                now,
            ),
        )

        conn.commit()

    return True

def purge_deleted_records_older_than(days: int = 30) -> int:
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT id
            FROM finance_records
            WHERE status = 'deleted'
              AND deleted_at IS NOT NULL
              AND DATETIME(deleted_at) <= DATETIME('now', '-{int(days)} day')
            """
        ).fetchall()

    record_ids = [row["id"] for row in rows]
    if not record_ids:
        return 0

    purged_count = 0

    for record_id in record_ids:
        record = get_record_by_id(record_id)
        if not record:
            continue

        attachments = list_attachments_for_record(record_id)

        with get_connection() as conn:
            conn.execute(
                "DELETE FROM finance_attachments WHERE finance_record_id = ?",
                (record_id,),
            )
            conn.execute(
                "DELETE FROM finance_record_history WHERE finance_record_id = ?",
                (record_id,),
            )
            conn.execute(
                "DELETE FROM finance_records WHERE id = ?",
                (record_id,),
            )
            conn.commit()

        for attachment in attachments:
            file_path = FINANCE_UPLOADS_DIR / attachment["stored_name"]
            if file_path.exists():
                try:
                    file_path.unlink()
                except OSError:
                    pass

        purged_count += 1

    return purged_count

def get_vendor_department_context(
    *,
    user_departments: list[dict],
    vendor_id: int,
    posted_department_name: str | None = None,
) -> str | None:
    posted_department_name = normalize_text(posted_department_name)
    if posted_department_name:
        return posted_department_name

    related_records = list_records_for_vendor(vendor_id)
    if related_records:
        return related_records[0]["department_name"]

    if user_departments:
        return user_departments[0]["department_name"]

    return None

def list_import_profiles() -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_import_profiles
            WHERE SUBSTR(profile_name, 1, 5) <> '_run_'
            ORDER BY profile_name COLLATE NOCASE
            """
        ).fetchall()

    return [dict(row) for row in rows]


def list_import_runs(limit: int = 25) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_import_runs
            ORDER BY started_at DESC, id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()

    return [dict(row) for row in rows]


def list_import_sources() -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_import_sources
            ORDER BY source_name COLLATE NOCASE
            """
        ).fetchall()

    return [dict(row) for row in rows]

def create_import_run(
    *,
    import_type: str,
    source_type: str,
    profile_id: int | None = None,
    original_filename: str | None = None,
    stored_filename: str | None = None,
    status: str = "pending",
    started_by_user_id: int | None = None,
) -> int:
    now = utc_now_iso()

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO finance_import_runs (
                import_type,
                source_type,
                profile_id,
                original_filename,
                stored_filename,
                status,
                started_by_user_id,
                started_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalize_text(import_type),
                normalize_text(source_type),
                profile_id,
                normalize_text(original_filename),
                normalize_text(stored_filename),
                normalize_text(status) or "pending",
                started_by_user_id,
                now,
            ),
        )
        conn.commit()
        return cursor.lastrowid
    
def get_import_run_by_id(run_id: int) -> dict | None:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_import_runs
            WHERE id = ?
            """,
            (run_id,),
        ).fetchone()

    return dict(row) if row else None

def update_import_run_status(
    run_id: int,
    *,
    status: str,
    total_rows: int | None = None,
    run_notes: str | None = None,
    completed: bool = False,
):
    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_import_runs
            SET
                status = ?,
                total_rows = COALESCE(?, total_rows),
                run_notes = COALESCE(?, run_notes),
                completed_at = CASE WHEN ? THEN ? ELSE completed_at END
            WHERE id = ?
            """,
            (
                normalize_text(status) or "pending",
                total_rows,
                normalize_text(run_notes),
                1 if completed else 0,
                now,
                run_id,
            ),
        )
        conn.commit()

def save_import_upload(original_filename: str, file_bytes: bytes) -> str:
    if not original_filename:
        raise ValueError("original_filename is required")
    if not file_bytes:
        raise ValueError("file_bytes is required")

    ensure_finance_imports_dir()

    ext = Path(original_filename).suffix.lower()
    stored_name = f"{secrets.token_hex(16)}{ext}"
    stored_path = FINANCE_IMPORTS_DIR / stored_name

    with open(stored_path, "wb") as f:
        f.write(file_bytes)

    return stored_name

def read_import_headers(stored_filename: str) -> list[str]:
    if not stored_filename:
        raise ValueError("stored_filename is required")

    file_path = FINANCE_IMPORTS_DIR / stored_filename
    if not file_path.exists():
        raise FileNotFoundError(f"Import file not found: {stored_filename}")

    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        with open(file_path, "rb") as raw_file:
            text_file = TextIOWrapper(raw_file, encoding="utf-8-sig", newline="")
            reader = csv.reader(text_file)
            headers = next(reader, [])
            return [str(h).strip() for h in headers if str(h).strip()]

    if suffix == ".xlsx":
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = workbook.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        return [str(h).strip() for h in first_row if h is not None and str(h).strip()]

    raise ValueError("Unsupported file type. Only CSV and XLSX are supported.")

def get_import_target_fields(import_type: str) -> list[dict]:
    import_type = (import_type or "").strip().lower()

    if import_type == "transactions":
        return [
            {"field_name": "fund", "label": "Fund", "required": False},
            {"field_name": "budget_unit", "label": "Budget Unit", "required": False},
            {"field_name": "account_code", "label": "Account Code", "required": False},
            {"field_name": "purchase_date", "label": "Date", "required": False},
            {"field_name": "po_number", "label": "PO Number", "required": False},
            {"field_name": "vendor_name", "label": "Vendor", "required": False},
            {"field_name": "expenditure_amount", "label": "Expenditures", "required": False},
            {"field_name": "encumbrance_amount", "label": "Encumbrances", "required": False},
            {"field_name": "description", "label": "Description", "required": False},
            {"field_name": "cumulative_balance", "label": "Cumulative Balance", "required": False},
        ]

    if import_type == "vendors":
        return [
            {"field_name": "vendor_name", "label": "Vendor Name", "required": True},
            {"field_name": "vendor_code", "label": "Vendor Code", "required": False},
            {"field_name": "website", "label": "Website", "required": False},
            {"field_name": "main_phone", "label": "Main Phone", "required": False},
            {"field_name": "billing_email", "label": "Billing Email", "required": False},
            {"field_name": "support_email", "label": "Support Email", "required": False},
            {"field_name": "sales_contact_name", "label": "Sales Contact Name", "required": False},
            {"field_name": "sales_contact_email", "label": "Sales Contact Email", "required": False},
            {"field_name": "notes", "label": "Notes", "required": False},
        ]

    return [
        {"field_name": "title", "label": "Item / Service Name", "required": True},
        {"field_name": "record_type", "label": "Record Type", "required": False},
        {"field_name": "vendor_name", "label": "Vendor Name", "required": False},
        {"field_name": "vendor_code", "label": "Vendor Code", "required": False},
        {"field_name": "category_name", "label": "Category", "required": False},
        {"field_name": "account_code", "label": "Account Code", "required": False},
        {"field_name": "po_number", "label": "PO Number", "required": False},
        {"field_name": "purchase_date", "label": "Purchase Date", "required": False},
        {"field_name": "service_start_date", "label": "Service Start Date", "required": False},
        {"field_name": "term_length", "label": "Term Length", "required": False},
        {"field_name": "term_unit", "label": "Term Unit", "required": False},
        {"field_name": "expiration_date", "label": "Expiration Date", "required": False},
        {"field_name": "renewal_date", "label": "Renewal Date", "required": False},
        {"field_name": "notify_days_before", "label": "Notify Days Before", "required": False},
        {"field_name": "notification_recipients", "label": "Notification Recipients", "required": False},
        {"field_name": "status", "label": "Status", "required": False},
        {"field_name": "cost", "label": "Cost", "required": False},
        {"field_name": "notes", "label": "Notes", "required": False},
    ]

def replace_import_profile_fields(profile_id: int, field_mappings: list[dict]):
    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            "DELETE FROM finance_import_profile_fields WHERE profile_id = ?",
            (profile_id,),
        )

        for item in field_mappings:
            ignore_field = 1 if item.get("ignore_field") else 0
            source_column_name = None if ignore_field else normalize_text(item.get("source_column_name"))

            conn.execute(
                """
                INSERT INTO finance_import_profile_fields (
                    profile_id,
                    source_column_name,
                    target_field_name,
                    transform_rule,
                    default_value,
                    required,
                    ignore_field,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    profile_id,
                    source_column_name,
                    normalize_text(item.get("target_field_name")),
                    normalize_text(item.get("transform_rule")),
                    normalize_text(item.get("default_value")),
                    1 if item.get("required") else 0,
                    ignore_field,
                    now,
                    now,
                ),
            )

        conn.commit()

def build_friendly_import_title(source_row: dict, row_number: int | None = None) -> str:
    description = normalize_text(source_row.get("DESCRIPTION"))
    vendor = normalize_text(source_row.get("VENDOR"))
    account = normalize_text(source_row.get("ACCOUNT"))
    po_number = normalize_text(source_row.get("PURCHASE O"))
    purchase_date = normalize_text(source_row.get("DATE"))

    if description:
        upper_description = description.upper()
        if not upper_description.startswith("TOTAL ") and upper_description not in {
            "BEGINNING BALANCE",
            "POSTED FROM BUDGET SYSTEM",
        }:
            return description.title()

    if vendor and po_number:
        return f"{vendor.title()} - {po_number}"

    if vendor and account:
        return f"{vendor.title()} - Account {account}"

    if vendor:
        return vendor.title()

    if account and purchase_date:
        return f"Account {account} - {purchase_date}"

    if account:
        return f"Account {account}"

    if row_number:
        return f"Finance Import Row {row_number}"

    return "Imported Finance Item"

def split_efinance_vendor(value: str | None) -> tuple[str | None, str | None]:
    value = normalize_text(value)
    if not value:
        return None, None

    parts = value.split(maxsplit=1)

    if len(parts) == 2 and parts[0].isdigit():
        return normalize_vendor_name(parts[1]), parts[0]

    return normalize_vendor_name(value), None


def parse_money_decimal(value) -> Decimal:
    return parse_cost_to_decimal(value)


def money_to_string(value) -> str:
    amount = parse_money_decimal(value)
    return str(amount.quantize(Decimal("0.01")))


def should_skip_efinance_transaction(mapped: dict, source_row: dict) -> bool:
    description = (normalize_text(mapped.get("description")) or "").upper()
    fund = (normalize_text(mapped.get("fund")) or "").upper()

    expenditure = parse_money_decimal(mapped.get("expenditure_amount"))
    encumbrance = parse_money_decimal(mapped.get("encumbrance_amount"))
    balance = parse_money_decimal(mapped.get("cumulative_balance"))

    if description in {"BEGINNING BALANCE", "POSTED FROM BUDGET SYSTEM"}:
        return True

    if fund.startswith("TOTAL "):
        return True

    if description.startswith("TOTAL "):
        return True

    if not description and expenditure == 0 and encumbrance == 0 and balance == 0:
        return True

    # subtotal/balance-only rows
    if not description and not mapped.get("vendor_name") and not mapped.get("po_number"):
        return True

    return False


def classify_transaction_type(mapped: dict) -> str:
    description = (normalize_text(mapped.get("description")) or "").upper()
    expenditure = parse_money_decimal(mapped.get("expenditure_amount"))
    encumbrance = parse_money_decimal(mapped.get("encumbrance_amount"))

    if "CHANGE ORDER" in description:
        return "change_order"

    if "BLANKET PO" in description:
        return "blanket_po"

    if "TAX" in description:
        return "tax_fee"

    if expenditure < 0:
        return "credit_refund"

    if expenditure == 0 and encumbrance > 0:
        return "encumbrance"

    if encumbrance < 0 and expenditure == 0:
        return "encumbrance_release"

    if expenditure > 0 and encumbrance < 0:
        return "purchase"

    if expenditure > 0:
        return "purchase"

    if encumbrance != 0:
        return "encumbrance"

    return "other"


def suggest_record_type_for_transaction(mapped: dict) -> str | None:
    description = (normalize_text(mapped.get("description")) or "").upper()

    renewal_words = [
        "ANNUAL",
        "RENEWAL",
        "SUBSCRIPTION",
        "LICENSE",
        "SOFTWARE",
        "SUPPORT",
        "SITE LICENSE",
        "SERVICE AGREEMENT",
    ]

    if any(word in description for word in renewal_words):
        return "renewal"

    if "COPIER SERVICE AGREEMENT" in description:
        return "service_contract"

    if "BLANKET PO" in description:
        return "blanket_po"

    if "CHANGE ORDER" in description:
        return "change_order"

    return None


def build_transaction_title(mapped: dict, source_row: dict, row_number: int | None = None) -> str:
    description = normalize_text(mapped.get("description"))
    vendor_name = normalize_text(mapped.get("vendor_name"))
    po_number = normalize_text(mapped.get("po_number"))
    account_code = normalize_text(mapped.get("account_code"))

    if description:
        return description.title()

    if vendor_name and po_number:
        return f"{vendor_name} - PO {po_number}"

    if vendor_name and account_code:
        return f"{vendor_name} - Account {account_code}"

    if account_code:
        return f"Account {account_code}"

    if row_number:
        return f"Finance Transaction Row {row_number}"

    return "Finance Transaction"


def create_finance_transaction(
    *,
    department_name: str,
    import_run_id: int | None,
    source_type: str | None,
    source_row_number: int | None,
    transaction_type: str,
    review_status: str,
    title: str,
    description: str | None,
    vendor_id: int | None,
    vendor_code: str | None,
    vendor_name: str | None,
    fund: str | None,
    budget_unit: str | None,
    account_code: str | None,
    po_number: str | None,
    purchase_date: str | None,
    expenditure_amount: str | None,
    encumbrance_amount: str | None,
    cumulative_balance: str | None,
    suggested_record_type: str | None,
    is_promotable: bool,
    raw_json: str | None,
) -> int:
    now = utc_now_iso()

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO finance_transactions (
                department_name,
                import_run_id,
                source_type,
                source_row_number,
                transaction_type,
                review_status,
                title,
                description,
                vendor_id,
                vendor_code,
                vendor_name,
                fund,
                budget_unit,
                account_code,
                po_number,
                purchase_date,
                expenditure_amount,
                encumbrance_amount,
                cumulative_balance,
                suggested_record_type,
                is_promotable,
                raw_json,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalize_text(department_name),
                import_run_id,
                normalize_text(source_type),
                source_row_number,
                normalize_text(transaction_type),
                normalize_text(review_status) or "needs_review",
                normalize_text(title),
                normalize_text(description),
                vendor_id,
                normalize_text(vendor_code),
                normalize_text(vendor_name),
                normalize_text(fund),
                normalize_text(budget_unit),
                normalize_text(account_code),
                normalize_text(po_number),
                normalize_text(purchase_date),
                normalize_text(expenditure_amount),
                normalize_text(encumbrance_amount),
                normalize_text(cumulative_balance),
                normalize_text(suggested_record_type),
                1 if is_promotable else 0,
                raw_json,
                now,
                now,
            ),
        )
        conn.commit()
        return cursor.lastrowid


def list_transactions_for_department(
    department_name: str,
    review_status: str | None = None,
    transaction_type: str | None = None,
    vendor_q: str | None = None,
    page: int = 1,
    per_page: int = 100,
) -> dict:
    department_name = normalize_text(department_name)
    review_status = normalize_text(review_status)
    transaction_type = normalize_text(transaction_type)
    vendor_q = normalize_text(vendor_q)

    page = max(int(page or 1), 1)
    per_page = max(min(int(per_page or 100), 250), 25)
    offset = (page - 1) * per_page

    where = ["department_name = ?"]
    params = [department_name]

    if review_status:
        where.append("review_status = ?")
        params.append(review_status)

    if transaction_type:
        where.append("transaction_type = ?")
        params.append(transaction_type)

    if vendor_q:
        where.append("LOWER(vendor_name) LIKE LOWER(?)")
        params.append(f"%{vendor_q}%")

    where_sql = " AND ".join(where)

    with get_connection() as conn:
        total = conn.execute(
            f"""
            SELECT COUNT(*) AS count
            FROM finance_transactions
            WHERE {where_sql}
            """,
            params,
        ).fetchone()["count"]

        rows = conn.execute(
            f"""
            SELECT *
            FROM finance_transactions
            WHERE {where_sql}
            ORDER BY
                CASE WHEN purchase_date IS NULL OR purchase_date = '' THEN 1 ELSE 0 END,
                purchase_date DESC,
                id DESC
            LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()

    total_pages = max((total + per_page - 1) // per_page, 1)

    return {
        "rows": [dict(row) for row in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": page - 1,
        "next_page": page + 1,
    }

def get_transaction_by_id(transaction_id: int) -> dict | None:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_transactions
            WHERE id = ?
            """,
            (transaction_id,),
        ).fetchone()

    return dict(row) if row else None


def mark_transaction_promoted(transaction_id: int, record_id: int):
    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_transactions
            SET review_status = 'promoted',
                promoted_record_id = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (record_id, now, transaction_id),
        )
        conn.commit()

def bulk_promote_transactions_to_records(
    *,
    transaction_ids: list[int],
    department_name: str,
    created_by_user_id: int | None = None,
) -> dict:
    department_name = normalize_text(department_name)

    if not department_name:
        raise ValueError("department_name is required")

    if not transaction_ids:
        return {
            "created": 0,
            "skipped": 0,
            "errors": [],
        }

    created = 0
    skipped = 0
    errors = []

    for transaction_id in transaction_ids:
        try:
            transaction = get_transaction_by_id(transaction_id)

            if not transaction:
                skipped += 1
                errors.append(f"Transaction {transaction_id} was not found.")
                continue

            if transaction.get("department_name") != department_name:
                skipped += 1
                errors.append(f"Transaction {transaction_id} does not belong to this department.")
                continue

            if (transaction.get("review_status") or "").lower() == "promoted":
                skipped += 1
                continue

            title = (
                normalize_text(transaction.get("title"))
                or normalize_text(transaction.get("description"))
                or f"Transaction {transaction_id}"
            )

            expenditure_amount = normalize_text(transaction.get("expenditure_amount"))
            encumbrance_amount = normalize_text(transaction.get("encumbrance_amount"))

            cost = expenditure_amount or encumbrance_amount or "0.00"

            transaction_type = (transaction.get("transaction_type") or "").strip().lower()

            if transaction_type in {"purchase", "credit_refund", "tax_fee"}:
                record_type = "purchase"
            else:
                record_type = transaction.get("suggested_record_type") or "renewal"

            record_id = create_record(
                record_type=record_type,
                title=title,
                department_name=department_name,
                vendor_id=transaction.get("vendor_id"),
                category_id=None,
                account_code=transaction.get("account_code"),
                po_number=transaction.get("po_number"),
                purchase_date=transaction.get("purchase_date"),
                service_start_date="",
                use_purchase_date_as_start=True,
                term_length=None,
                term_unit="",
                expiration_date="",
                renewal_date="",
                notify_days_before=30,
                notification_recipients="",
                status="active",
                cost=cost,
                notes=(
                    f"Bulk promoted from Finance Transaction #{transaction_id}\n\n"
                    f"Description: {transaction.get('description') or ''}\n"
                    f"Transaction Type: {transaction.get('transaction_type') or ''}\n"
                    f"Encumbrance Amount: {transaction.get('encumbrance_amount') or ''}\n"
                    f"Cumulative Balance: {transaction.get('cumulative_balance') or ''}"
                ),
                created_by_user_id=created_by_user_id,
            )

            mark_transaction_promoted(transaction_id, record_id)
            created += 1

        except Exception as exc:
            skipped += 1
            errors.append(f"Transaction {transaction_id}: {exc}")

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }

def create_import_profile(
    *,
    profile_name: str,
    source_type: str,
    target_area: str,
    description: str | None = None,
    created_by_user_id: int | None = None,
) -> int:
    profile_name = normalize_text(profile_name)
    source_type = normalize_text(source_type)
    target_area = normalize_text(target_area)

    if not profile_name:
        raise ValueError("profile_name is required")
    if not source_type:
        raise ValueError("source_type is required")
    if not target_area:
        raise ValueError("target_area is required")

    now = utc_now_iso()

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO finance_import_profiles (
                profile_name,
                source_type,
                target_area,
                is_active,
                description,
                created_by_user_id,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, 1, ?, ?, ?, ?)
            """,
            (
                profile_name,
                source_type,
                target_area,
                normalize_text(description),
                created_by_user_id,
                now,
                now,
            ),
        )
        conn.commit()
        return cursor.lastrowid
    
def get_vendor_template_headers() -> list[str]:
    return [
        "vendor_name",
        "vendor_code",
        "website",
        "main_phone",
        "billing_email",
        "support_email",
        "sales_contact_name",
        "sales_contact_email",
        "notes",
    ]

def get_record_template_headers() -> list[str]:
    return [
        "title",
        "record_type",
        "vendor_name",
        "vendor_code",
        "category_name",
        "account_code",
        "po_number",
        "purchase_date",
        "service_start_date",
        "term_length",
        "term_unit",
        "expiration_date",
        "renewal_date",
        "notify_days_before",
        "notification_recipients",
        "status",
        "cost",
        "notes",
    ]

def get_import_profile_by_name(profile_name: str) -> dict | None:
    profile_name = normalize_text(profile_name)
    if not profile_name:
        return None

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_import_profiles
            WHERE profile_name = ?
            """,
            (profile_name,),
        ).fetchone()

    return dict(row) if row else None

def get_import_profile_fields(profile_id: int) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_import_profile_fields
            WHERE profile_id = ?
            ORDER BY id ASC
            """,
            (profile_id,),
        ).fetchall()

    return [dict(row) for row in rows]


def set_import_run_profile(run_id: int, profile_id: int):
    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_import_runs
            SET profile_id = ?
            WHERE id = ?
            """,
            (profile_id, run_id),
        )
        conn.commit()


def log_import_run_error(
    *,
    run_id: int,
    row_number: int | None,
    source_identifier: str | None,
    error_message: str,
):
    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO finance_import_run_errors (
                run_id,
                row_number,
                source_identifier,
                error_message,
                created_at
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                run_id,
                row_number,
                normalize_text(source_identifier),
                error_message,
                now,
            ),
        )
        conn.commit()


def list_import_run_errors(run_id: int) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_import_run_errors
            WHERE run_id = ?
            ORDER BY id ASC
            """,
            (run_id,),
        ).fetchall()

    return [dict(row) for row in rows]


def update_import_run_results(
    run_id: int,
    *,
    status: str,
    total_rows: int,
    created_rows: int,
    updated_rows: int,
    skipped_rows: int,
    error_rows: int,
    run_notes: str | None = None,
    completed: bool = False,
):
    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_import_runs
            SET
                status = ?,
                total_rows = ?,
                created_rows = ?,
                updated_rows = ?,
                skipped_rows = ?,
                error_rows = ?,
                run_notes = ?,
                completed_at = CASE WHEN ? THEN ? ELSE completed_at END
            WHERE id = ?
            """,
            (
                normalize_text(status) or "pending",
                total_rows,
                created_rows,
                updated_rows,
                skipped_rows,
                error_rows,
                normalize_text(run_notes),
                1 if completed else 0,
                now,
                run_id,
            ),
        )
        conn.commit()


def read_import_rows(stored_filename: str) -> list[dict]:
    if not stored_filename:
        raise ValueError("stored_filename is required")

    file_path = FINANCE_IMPORTS_DIR / stored_filename
    if not file_path.exists():
        raise FileNotFoundError(f"Import file not found: {stored_filename}")

    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        with open(file_path, "rb") as raw_file:
            text_file = TextIOWrapper(raw_file, encoding="utf-8-sig", newline="")
            reader = csv.DictReader(text_file)
            return [
                {str(k).strip(): ("" if v is None else str(v).strip()) for k, v in row.items()}
                for row in reader
            ]

    if suffix == ".xlsx":
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = workbook.active

        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []

        header_names = [str(h).strip() if h is not None else "" for h in headers]

        results = []
        for row in rows_iter:
            item = {}
            for idx, header in enumerate(header_names):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                item[header] = "" if value is None else str(value).strip()
            results.append(item)

        return results

    raise ValueError("Unsupported file type. Only CSV and XLSX are supported.")

def apply_import_mapping(row: dict, mappings: list[dict], row_number: int | None = None) -> dict:
    mapped = {}

    for mapping in mappings:
        if mapping["ignore_field"]:
            continue

        target_field = mapping["target_field_name"]
        source_column = mapping["source_column_name"]
        transform_rule = normalize_text(mapping.get("transform_rule"))
        default_value = normalize_text(mapping.get("default_value"))

        if transform_rule == "friendly_title":
            value = build_friendly_import_title(row, row_number)

        elif transform_rule == "vendor_split":
            vendor_name, vendor_code = split_efinance_vendor(row.get(source_column))
            mapped["vendor_name"] = vendor_name
            mapped["vendor_code"] = vendor_code
            continue

        elif transform_rule == "money":
            value = money_to_string(row.get(source_column))

        elif source_column:
            value = normalize_text(row.get(source_column))

        else:
            value = None

        if not value and default_value:
            value = default_value

        mapped[target_field] = value

    return mapped

def preview_import_rows(
    *,
    run_id: int,
    profile_id: int,
    limit: int = 10,
) -> list[dict]:
    run = get_import_run_by_id(run_id)
    if not run:
        raise ValueError("Import run not found")

    rows = read_import_rows(run["stored_filename"])
    mappings = get_import_profile_fields(profile_id)

    target_to_source = {}
    for mapping in mappings:
        if mapping["ignore_field"]:
            continue
        if not mapping["source_column_name"]:
            continue
        target_to_source[mapping["target_field_name"]] = mapping["source_column_name"]

    preview = []
    for row in rows[:limit]:
        mapped = {}
        for target_field, source_column in target_to_source.items():
            mapped[target_field] = row.get(source_column, "")
        preview.append(mapped)

    return preview


def find_vendor_for_import(vendor_name: str | None = None, vendor_code: str | None = None) -> dict | None:
    vendor_name = normalize_text(vendor_name)
    vendor_code = normalize_text(vendor_code)

    with get_connection() as conn:
        if vendor_code:
            row = conn.execute(
                """
                SELECT *
                FROM finance_vendors
                WHERE vendor_code = ?
                LIMIT 1
                """,
                (vendor_code,),
            ).fetchone()
            if row:
                return dict(row)

        if vendor_name:
            row = conn.execute(
                """
                SELECT *
                FROM finance_vendors
                WHERE LOWER(TRIM(vendor_name)) = LOWER(TRIM(?))
                LIMIT 1
                """,
                (vendor_name,),
            ).fetchone()
            if row:
                return dict(row)

    return None


def get_or_create_vendor_for_import(
    *,
    vendor_name: str | None = None,
    vendor_code: str | None = None,
) -> tuple[int | None, bool]:
    vendor_name = normalize_vendor_name(vendor_name)
    vendor_code = normalize_text(vendor_code)

    if not vendor_name and not vendor_code:
        return None, False

    existing = find_vendor_for_import(vendor_name=vendor_name, vendor_code=vendor_code)
    if existing:
        return existing["id"], False

    if vendor_name:
        vendor_id = create_vendor(
            vendor_name=vendor_name,
            vendor_code=vendor_code,
        )
        return vendor_id, True

    return None, False


def execute_records_import(
    *,
    run_id: int,
    profile_id: int,
    default_department_name: str | None = None,
    created_by_user_id: int | None = None,
) -> dict:
    run = get_import_run_by_id(run_id)
    if not run:
        raise ValueError("Import run not found")

    rows = read_import_rows(run["stored_filename"])
    mappings = get_import_profile_fields(profile_id)

    total_rows = len(rows)
    created_rows = 0
    skipped_rows = 0
    error_rows = 0
    vendors_created = 0
    notifications_sent = 0

    update_import_run_results(
        run_id,
        status="running",
        total_rows=total_rows,
        created_rows=0,
        updated_rows=0,
        skipped_rows=0,
        error_rows=0,
        run_notes="Import execution started.",
        completed=False,
    )

    for index, row in enumerate(rows, start=2):
        try:
            mapped = apply_import_mapping(row, mappings, index)

            if default_department_name:
                mapped["department_name"] = normalize_text(default_department_name)

            missing_required = []

            for mapping in mappings:
                if not mapping["required"]:
                    continue

                if mapping["ignore_field"]:
                    continue

                target_field = mapping["target_field_name"]

                if not normalize_text(mapped.get(target_field)):
                    missing_required.append(target_field)

            if missing_required:
                skipped_rows += 1
                log_import_run_error(
                    run_id=run_id,
                    row_number=index,
                    source_identifier=(
                        mapped.get("title")
                        or row.get(next(iter(row.keys()), ""), "")
                        or f"Row {index}"
                    ),
                    error_message=f"Missing required mapped fields: {', '.join(missing_required)}",
                )
                continue

            vendor_id, vendor_was_created = get_or_create_vendor_for_import(
                vendor_name=mapped.get("vendor_name"),
                vendor_code=mapped.get("vendor_code"),
            )

            if vendor_was_created:
                vendors_created += 1

            category_name = normalize_text(mapped.get("category_name"))
            category = find_category_for_import(category_name)

            if category_name and not category:
                suggestion = find_category_suggestion_for_import(category_name)

                if suggestion:
                    confidence_percent = int(round(suggestion["confidence_score"] * 100))
                    raise ValueError(
                        f"Category not mapped: {category_name}. "
                        f"Suggested match: {suggestion['suggested_category_name']} "
                        f"({confidence_percent}% confidence)."
                    )

                raise ValueError(f"Category not mapped: {category_name}")

            category_id = category["id"] if category else None

            term_length = None
            if mapped.get("term_length"):
                try:
                    term_length = int(str(mapped["term_length"]).strip())
                except ValueError:
                    term_length = None

            notify_days_before = 30
            if mapped.get("notify_days_before"):
                try:
                    notify_days_before = int(str(mapped["notify_days_before"]).strip())
                except ValueError:
                    notify_days_before = 30

            title = normalize_text(mapped.get("title")) or "Imported Finance Item"

            record_id = create_record(
                record_type=mapped.get("record_type") or "renewal",
                title=title,
                department_name=mapped.get("department_name") or "",
                vendor_id=vendor_id,
                category_id=category_id,
                account_code=mapped.get("account_code"),
                po_number=mapped.get("po_number"),
                purchase_date=mapped.get("purchase_date"),
                service_start_date=mapped.get("service_start_date"),
                use_purchase_date_as_start=not bool(mapped.get("service_start_date")),
                term_length=term_length,
                term_unit=mapped.get("term_unit"),
                expiration_date=mapped.get("expiration_date"),
                renewal_date=mapped.get("renewal_date"),
                notify_days_before=notify_days_before,
                notification_recipients=mapped.get("notification_recipients"),
                status=mapped.get("status") or "active",
                cost=mapped.get("cost"),
                notes=mapped.get("notes"),
                created_by_user_id=created_by_user_id,
            )

            sent_now, _ = maybe_send_renewal_notification_for_record(
                record_id,
                changed_by_user_id=created_by_user_id,
            )
            if sent_now:
                notifications_sent += 1

            created_rows += 1

        except Exception as exc:
            error_rows += 1
            log_import_run_error(
                run_id=run_id,
                row_number=index,
                source_identifier=row.get(next(iter(row.keys()), ""), "") if row else None,
                error_message=str(exc),
            )

    run_notes = (
        f"Import completed. Created records: {created_rows}, "
        f"Created vendors: {vendors_created}, "
        f"Notifications sent: {notifications_sent}, "
        f"Skipped: {skipped_rows}, Errors: {error_rows}."
    )

    final_status = "completed"
    if error_rows and not created_rows:
        final_status = "completed_with_errors"
    elif error_rows:
        final_status = "completed_with_warnings"

    update_import_run_results(
        run_id,
        status=final_status,
        total_rows=total_rows,
        created_rows=created_rows,
        updated_rows=0,
        skipped_rows=skipped_rows,
        error_rows=error_rows,
        run_notes=run_notes,
        completed=True,
    )

    return {
        "total_rows": total_rows,
        "created_rows": created_rows,
        "updated_rows": 0,
        "skipped_rows": skipped_rows,
        "error_rows": error_rows,
        "vendors_created": vendors_created,
        "notifications_sent": notifications_sent,
        "status": final_status,
        "run_notes": run_notes,
    }

def get_import_run_field_map(run_id: int) -> dict[str, str]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT target_field_name, source_column_name
            FROM finance_import_profile_fields
            WHERE profile_id = (
                SELECT profile_id
                FROM finance_import_runs
                WHERE id = ?
            )
              AND ignore_field = 0
              AND source_column_name IS NOT NULL
            """,
            (run_id,),
        ).fetchall()

    return {
        row["target_field_name"]: row["source_column_name"]
        for row in rows
    }

def get_import_profile_field_map(profile_id: int) -> dict[str, str]:
    rows = get_import_profile_fields(profile_id)
    result = {}

    for row in rows:
        if row.get("ignore_field"):
            continue

        target_field = row.get("target_field_name")
        source_column = row.get("source_column_name")

        if target_field and source_column:
            result[target_field] = source_column

    return result

def clear_import_run_errors(run_id: int):
    with get_connection() as conn:
        conn.execute(
            """
            DELETE FROM finance_import_run_errors
            WHERE run_id = ?
            """,
            (run_id,),
        )
        conn.commit()


def validate_records_import(
    *,
    run_id: int,
    profile_id: int,
    default_department_name: str | None = None,
    preview_limit: int = 20,
) -> dict:
    run = get_import_run_by_id(run_id)
    if not run:
        raise ValueError("Import run not found")

    rows = read_import_rows(run["stored_filename"])
    mappings = get_import_profile_fields(profile_id)

    clear_import_run_errors(run_id)

    preview_rows = []
    errors = []
    category_suggestions = {}
    valid_rows = 0
    skipped_rows = 0
    vendors_to_create = 0

    for index, row in enumerate(rows, start=2):
        mapped = apply_import_mapping(row, mappings, index)

        if default_department_name:
            mapped["department_name"] = normalize_text(default_department_name)

        row_errors = []

        for mapping in mappings:
            if not mapping["required"]:
                continue

            if mapping["ignore_field"]:
                continue

            target_field = mapping["target_field_name"]

            if not normalize_text(mapped.get(target_field)):
                row_errors.append(f"Missing required field: {target_field}")

        vendor_name = normalize_text(mapped.get("vendor_name"))
        vendor_code = normalize_text(mapped.get("vendor_code"))

        if vendor_name or vendor_code:
            existing_vendor = find_vendor_for_import(
                vendor_name=vendor_name,
                vendor_code=vendor_code,
            )

            if not existing_vendor and vendor_name:
                vendors_to_create += 1

        category_name = normalize_text(mapped.get("category_name"))

        if category_name:
            existing_category = find_category_for_import(category_name)

            if existing_category:
                mapped["_category_match_status"] = "matched"
                mapped["_category_match_label"] = existing_category["category_name"]

            else:
                suggestion = find_category_suggestion_for_import(category_name)

                if suggestion:
                    category_suggestions[suggestion["id"]] = suggestion

                    confidence_percent = int(
                        round(float(suggestion["confidence_score"] or 0) * 100)
                    )

                    suggested_category_name = (
                        suggestion.get("suggested_category_name") or "Unknown"
                    )

                    mapped["_category_match_status"] = "suggested"
                    mapped["_category_match_label"] = (
                        f"{suggested_category_name} ({confidence_percent}% match)"
                    )

                    row_errors.append(
                        f"Category not mapped: {category_name}. "
                        f"Suggested match: {suggested_category_name} "
                        f"({confidence_percent}% confidence)."
                    )

                else:
                    mapped["_category_match_status"] = "unmatched"
                    mapped["_category_match_label"] = "No suggested match"
                    row_errors.append(f"Category not mapped: {category_name}")

        if mapped.get("term_length"):
            try:
                int(str(mapped["term_length"]).strip())
            except ValueError:
                row_errors.append(f"Invalid term_length: {mapped.get('term_length')}")

        if mapped.get("notify_days_before"):
            try:
                int(str(mapped["notify_days_before"]).strip())
            except ValueError:
                row_errors.append(
                    f"Invalid notify_days_before: {mapped.get('notify_days_before')}"
                )

        source_identifier = (
            mapped.get("title")
            or row.get(next(iter(row.keys()), ""), "")
            or f"Row {index}"
        )

        if row_errors:
            skipped_rows += 1
            error_message = "; ".join(row_errors)

            errors.append(
                {
                    "row_number": index,
                    "source_identifier": source_identifier,
                    "error_message": error_message,
                }
            )

            log_import_run_error(
                run_id=run_id,
                row_number=index,
                source_identifier=source_identifier,
                error_message=error_message,
            )

        else:
            valid_rows += 1

        if len(preview_rows) < preview_limit:
            preview_rows.append(mapped)

    return {
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "skipped_rows": skipped_rows,
        "vendors_to_create": vendors_to_create,
        "preview_rows": preview_rows,
        "errors": errors,
        "category_suggestions": list(category_suggestions.values()),
    }

def validate_vendors_import(
    *,
    run_id: int,
    profile_id: int,
    preview_limit: int = 20,
) -> dict:
    run = get_import_run_by_id(run_id)
    if not run:
        raise ValueError("Import run not found")

    rows = read_import_rows(run["stored_filename"])
    mappings = get_import_profile_fields(profile_id)

    clear_import_run_errors(run_id)

    preview_rows = []
    errors = []
    valid_rows = 0
    skipped_rows = 0
    vendors_to_create = 0
    vendors_to_update = 0

    for index, row in enumerate(rows, start=2):
        mapped = apply_import_mapping(row, mappings, index)
        row_errors = []

        vendor_name = normalize_vendor_name(mapped.get("vendor_name"))
        vendor_code = normalize_text(mapped.get("vendor_code"))

        if not vendor_name:
            row_errors.append("Missing required field: vendor_name")

        existing_vendor = None
        if vendor_name or vendor_code:
            existing_vendor = find_vendor_for_import(
                vendor_name=vendor_name,
                vendor_code=vendor_code,
            )

        source_identifier = (
            vendor_name
            or vendor_code
            or row.get(next(iter(row.keys()), ""), "")
            or f"Row {index}"
        )

        if row_errors:
            skipped_rows += 1
            error_message = "; ".join(row_errors)

            errors.append(
                {
                    "row_number": index,
                    "source_identifier": source_identifier,
                    "error_message": error_message,
                }
            )

            log_import_run_error(
                run_id=run_id,
                row_number=index,
                source_identifier=source_identifier,
                error_message=error_message,
            )
        else:
            valid_rows += 1

            if existing_vendor:
                vendors_to_update += 1
            else:
                vendors_to_create += 1

        if len(preview_rows) < preview_limit:
            preview_rows.append(
                {
                    "vendor_name": vendor_name or "",
                    "vendor_code": vendor_code or "",
                    "website": normalize_text(mapped.get("website")) or "",
                    "main_phone": normalize_text(mapped.get("main_phone")) or "",
                    "billing_email": normalize_text(mapped.get("billing_email")) or "",
                    "support_email": normalize_text(mapped.get("support_email")) or "",
                    "sales_contact_name": normalize_text(mapped.get("sales_contact_name")) or "",
                    "sales_contact_email": normalize_text(mapped.get("sales_contact_email")) or "",
                    "notes": normalize_text(mapped.get("notes")) or "",
                    "_import_action": "Update existing vendor" if existing_vendor else "Create new vendor",
                }
            )

    return {
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "skipped_rows": skipped_rows,
        "vendors_to_create": vendors_to_create,
        "vendors_to_update": vendors_to_update,
        "preview_rows": preview_rows,
        "errors": errors,
    }


def execute_vendors_import(
    *,
    run_id: int,
    profile_id: int,
    created_by_user_id: int | None = None,
) -> dict:
    run = get_import_run_by_id(run_id)
    if not run:
        raise ValueError("Import run not found")

    rows = read_import_rows(run["stored_filename"])
    mappings = get_import_profile_fields(profile_id)

    total_rows = len(rows)
    created_rows = 0
    updated_rows = 0
    skipped_rows = 0
    error_rows = 0

    update_import_run_results(
        run_id,
        status="running",
        total_rows=total_rows,
        created_rows=0,
        updated_rows=0,
        skipped_rows=0,
        error_rows=0,
        run_notes="Vendor import execution started.",
        completed=False,
    )

    for index, row in enumerate(rows, start=2):
        try:
            mapped = apply_import_mapping(row, mappings, index)

            vendor_name = normalize_vendor_name(mapped.get("vendor_name"))
            vendor_code = normalize_text(mapped.get("vendor_code"))

            if not vendor_name:
                skipped_rows += 1
                log_import_run_error(
                    run_id=run_id,
                    row_number=index,
                    source_identifier=vendor_code or f"Row {index}",
                    error_message="Missing required field: vendor_name",
                )
                continue

            existing_vendor = find_vendor_for_import(
                vendor_name=vendor_name,
                vendor_code=vendor_code,
            )

            if existing_vendor:
                update_vendor(
                    vendor_id=existing_vendor["id"],
                    vendor_name=vendor_name,
                    vendor_code=vendor_code,
                    website=mapped.get("website"),
                    main_phone=mapped.get("main_phone"),
                    billing_email=mapped.get("billing_email"),
                    support_email=mapped.get("support_email"),
                    sales_contact_name=mapped.get("sales_contact_name"),
                    sales_contact_email=mapped.get("sales_contact_email"),
                    status=existing_vendor.get("status") or "active",
                    notes=mapped.get("notes"),
                )
                updated_rows += 1
            else:
                create_vendor(
                    vendor_name=vendor_name,
                    vendor_code=vendor_code,
                    website=mapped.get("website"),
                    main_phone=mapped.get("main_phone"),
                    billing_email=mapped.get("billing_email"),
                    support_email=mapped.get("support_email"),
                    sales_contact_name=mapped.get("sales_contact_name"),
                    sales_contact_email=mapped.get("sales_contact_email"),
                    notes=mapped.get("notes"),
                )
                created_rows += 1

        except Exception as exc:
            error_rows += 1
            log_import_run_error(
                run_id=run_id,
                row_number=index,
                source_identifier=row.get(next(iter(row.keys()), ""), "") if row else None,
                error_message=str(exc),
            )

    run_notes = (
        f"Vendor import completed. Created vendors: {created_rows}, "
        f"Updated vendors: {updated_rows}, "
        f"Skipped: {skipped_rows}, Errors: {error_rows}."
    )

    final_status = "completed"
    if error_rows and not created_rows and not updated_rows:
        final_status = "completed_with_errors"
    elif error_rows:
        final_status = "completed_with_warnings"

    update_import_run_results(
        run_id,
        status=final_status,
        total_rows=total_rows,
        created_rows=created_rows,
        updated_rows=updated_rows,
        skipped_rows=skipped_rows,
        error_rows=error_rows,
        run_notes=run_notes,
        completed=True,
    )

    return {
        "total_rows": total_rows,
        "created_rows": created_rows,
        "updated_rows": updated_rows,
        "skipped_rows": skipped_rows,
        "error_rows": error_rows,
        "status": final_status,
        "run_notes": run_notes,
    }

def find_category_for_import(category_name: str | None) -> dict | None:
    category_name = normalize_text(category_name)
    if not category_name:
        return None

    normalized = normalize_category_lookup(category_name)

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_categories
            WHERE LOWER(TRIM(category_name)) = LOWER(TRIM(?))
              AND is_active = 1
            LIMIT 1
            """,
            (category_name,),
        ).fetchone()

        if row:
            category = dict(row)
            ensure_category_alias(category["id"], category_name)
            return category

        row = conn.execute(
            """
            SELECT c.*
            FROM finance_category_aliases a
            JOIN finance_categories c ON c.id = a.category_id
            WHERE a.normalized_alias = ?
              AND c.is_active = 1
            LIMIT 1
            """,
            (normalized,),
        ).fetchone()

        if row:
            return dict(row)

    return None

def normalize_import_header(value: str | None) -> str:
    value = normalize_text(value) or ""
    value = value.lower()
    value = value.replace("&", " and ")
    value = value.replace("-", "_")
    value = value.replace("/", "_")
    value = value.replace(" ", "_")

    while "__" in value:
        value = value.replace("__", "_")

    return value.strip("_")


def get_import_field_aliases(import_type: str) -> dict[str, list[str]]:
    import_type = (import_type or "").strip().lower()

    if import_type == "transactions":
        return {
            "fund": [
                "fund",
            ],
            "budget_unit": [
                "budget_unit",
                "budgetunit",
                "budget",
            ],
            "account_code": [
                "account",
                "account_code",
                "accountcode",
            ],
            "purchase_date": [
                "date",
                "purchase_date",
                "purchasedate",
            ],
            "po_number": [
                "purchase_o",
                "purchaseo",
                "purchase_order",
                "purchaseorder",
                "po",
                "po_number",
                "ponumber",
            ],
            "vendor_name": [
                "vendor",
                "vendor_name",
                "vendorname",
            ],
            "expenditure_amount": [
                "expenditures",
                "expenditure",
                "expense",
                "spent",
                "amount",
            ],
            "encumbrance_amount": [
                "encumbrances",
                "encumbrance",
                "encumbered",
            ],
            "description": [
                "description",
                "desc",
            ],
            "cumulative_balance": [
                "cumulative_balance",
                "cumulativebalance",
                "balance",
            ],
        }

    if import_type == "vendors":
        return {
            "vendor_name": ["vendor_name", "vendor", "name", "vendorname"],
            "vendor_code": ["vendor_code", "vendorcode", "code"],
            "website": ["website", "url", "site"],
            "main_phone": ["main_phone", "phone", "telephone"],
            "billing_email": ["billing_email", "billingemail"],
            "support_email": ["support_email", "supportemail"],
            "sales_contact_name": ["sales_contact_name", "salescontactname", "contact_name"],
            "sales_contact_email": ["sales_contact_email", "salescontactemail", "contact_email"],
            "notes": ["notes", "comments", "description"],
        }

    return {
        "title": ["title", "item_service_name", "item_name", "service_name", "name"],
        "record_type": ["record_type", "type"],
        "vendor_name": ["vendor_name", "vendor", "vendorname"],
        "vendor_code": ["vendor_code", "vendorcode"],
        "category_name": ["category_name", "category"],
        "account_code": ["account_code", "accountcode"],
        "po_number": ["po_number", "ponumber", "po"],
        "purchase_date": ["purchase_date", "purchasedate"],
        "service_start_date": ["service_start_date", "servicestartdate", "start_date", "startdate"],
        "term_length": ["term_length", "termlength"],
        "term_unit": ["term_unit", "termunit"],
        "expiration_date": ["expiration_date", "expirationdate", "end_date", "enddate"],
        "renewal_date": ["renewal_date", "renewaldate"],
        "notify_days_before": ["notify_days_before", "notifydaysbefore", "days_before"],
        "notification_recipients": ["notification_recipients", "notificationrecipients", "recipients"],
        "status": ["status"],
        "cost": ["cost", "amount", "price"],
        "notes": ["notes", "comments", "description"],
    }


def infer_import_field_map(source_headers: list[str], import_type: str) -> dict[str, str]:
    aliases = get_import_field_aliases(import_type)

    normalized_to_original = {}
    for header in source_headers:
        normalized_to_original[normalize_import_header(header)] = header

    inferred = {}

    for target_field, alias_list in aliases.items():
        for alias in alias_list:
            normalized_alias = normalize_import_header(alias)
            if normalized_alias in normalized_to_original:
                inferred[target_field] = normalized_to_original[normalized_alias]
                break

    return inferred

def parse_cost_to_decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    raw = str(value).strip()
    if not raw:
        return Decimal("0")

    cleaned = raw.replace("$", "").replace(",", "").strip()

    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return Decimal("0")

def list_active_budget_record_rows_for_department(department_name: str) -> list[dict]:
    department_name = normalize_text(department_name)

    if not department_name:
        return []

    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                r.*,
                v.vendor_name,
                c.category_name
            FROM finance_records r
            LEFT JOIN finance_vendors v ON v.id = r.vendor_id
            LEFT JOIN finance_categories c ON c.id = r.category_id
            WHERE r.department_name = ?
              AND r.status NOT IN ('archived', 'deleted')
            ORDER BY
                CASE WHEN r.renewal_date IS NULL THEN 1 ELSE 0 END,
                r.renewal_date ASC,
                r.title COLLATE NOCASE ASC
            """,
            (department_name,),
        ).fetchall()

    return [dict(row) for row in rows]


def get_budget_anchor_date(record: dict) -> date | None:
    if not isinstance(record, dict):
        return None

    for key in ("purchase_date", "renewal_date", "created_at"):
        value = record.get(key)

        if not value:
            continue

        if key == "created_at":
            try:
                return datetime.fromisoformat(value).date()
            except ValueError:
                continue

        parsed = _parse_iso_date(value)
        if parsed:
            return parsed

    return None


def list_budget_records_for_department(
    department_name: str,
    year: int | None = None,
    q: str | None = None,
) -> list[dict]:
    rows = list_active_budget_record_rows_for_department(department_name)
    q_norm = (q or "").strip().lower()

    selected_fiscal_year = get_budget_fiscal_year_by_year_number(year)

    filtered = []

    for row in rows:
        anchor_date = get_budget_anchor_date(row)
        record_fiscal_year = get_budget_fiscal_year_for_date(anchor_date)

        if selected_fiscal_year:
            if not record_fiscal_year:
                continue

            if int(record_fiscal_year["year_number"]) != int(selected_fiscal_year["year_number"]):
                continue

        if q_norm:
            haystack = " ".join(
                [
                    str(row.get("title") or ""),
                    str(row.get("vendor_name") or ""),
                    str(row.get("category_name") or ""),
                    str(row.get("account_code") or ""),
                    str(row.get("po_number") or ""),
                    str(row.get("notes") or ""),
                ]
            ).lower()

            if q_norm not in haystack:
                continue

        row = dict(row)
        row["_budget_anchor_date"] = anchor_date
        row["_budget_cost"] = parse_cost_to_decimal(row.get("cost"))
        row["_budget_fiscal_year"] = record_fiscal_year

        filtered.append(row)

    return filtered


def get_configured_budget_fiscal_years() -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM finance_fiscal_years
            ORDER BY year_number DESC
            """
        ).fetchall()

    return [dict(row) for row in rows]


def get_budget_fiscal_year_by_year_number(year_number: int | None) -> dict | None:
    if not year_number:
        return None

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_fiscal_years
            WHERE year_number = ?
            LIMIT 1
            """,
            (year_number,),
        ).fetchone()

    return dict(row) if row else None


def get_budget_fiscal_year_for_date(anchor_date: date | None) -> dict | None:
    if not anchor_date:
        return None

    anchor_value = anchor_date.isoformat()

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_fiscal_years
            WHERE DATE(?) BETWEEN DATE(start_date) AND DATE(end_date)
            ORDER BY year_number DESC
            LIMIT 1
            """,
            (anchor_value,),
        ).fetchone()

    return dict(row) if row else None


def get_budget_year_options_for_department(department_name: str) -> list[int]:
    fiscal_years = get_configured_budget_fiscal_years()
    return [int(item["year_number"]) for item in fiscal_years]

def get_budget_summary_for_department(
    department_name: str,
    year: int | None = None,
    q: str | None = None,
) -> dict:
    rows = list_budget_records_for_department(department_name, year=year, q=q)

    total_spent = sum((row["_budget_cost"] for row in rows), Decimal("0"))
    record_count = len(rows)
    average_spend = (total_spent / record_count) if record_count else Decimal("0")

    renewals_total = sum(
        (
            row["_budget_cost"]
            for row in rows
            if (row.get("record_type") or "").lower() == "renewal"
        ),
        Decimal("0"),
    )

    purchases_total = sum(
        (
            row["_budget_cost"]
            for row in rows
            if (row.get("record_type") or "").lower() == "purchase"
        ),
        Decimal("0"),
    )

    active_total = sum(
        (
            row["_budget_cost"]
            for row in rows
            if (row.get("status") or "").lower() == "active"
        ),
        Decimal("0"),
    )

    encumbrance_total = sum(
        (
            row["_budget_cost"]
            for row in rows
            if (row.get("status") or "").lower()
            in {"encumbered", "encumbrance", "pending", "pending_approval"}
        ),
        Decimal("0"),
    )

    budget_target = get_budget_target_for_department(department_name, year)
    total_budget = budget_target["total_budget"]
    remaining_budget = total_budget - total_spent

    percent_used = Decimal("0")

    if total_budget > 0:
        percent_used = (total_spent / total_budget) * Decimal("100")
    elif total_spent > 0:
        percent_used = Decimal("100")

    return {
        "total_budget": total_budget,
        "total_spent": total_spent,
        "remaining_budget": remaining_budget,
        "percent_used": percent_used.quantize(Decimal("0.1")),
        "record_count": record_count,
        "average_spend": average_spend,
        "renewals_total": renewals_total,
        "purchases_total": purchases_total,
        "active_total": active_total,
        "encumbrance_total": encumbrance_total,
        "budget_target": budget_target,
    }

def get_budget_dashboard_for_department(
    department_name: str,
    year: int | None = None,
    q: str | None = None,
) -> dict:
    return {
        "category": get_budget_breakdown_for_department(
            department_name=department_name,
            year=year,
            group_by="category",
            q=q,
        ),
        "vendor": get_budget_breakdown_for_department(
            department_name=department_name,
            year=year,
            group_by="vendor",
            q=q,
        ),
        "month": get_budget_breakdown_for_department(
            department_name=department_name,
            year=year,
            group_by="month",
            q=q,
        ),
        "record_type": get_budget_breakdown_for_department(
            department_name=department_name,
            year=year,
            group_by="record_type",
            q=q,
        ),
        "status": get_budget_breakdown_for_department(
            department_name=department_name,
            year=year,
            group_by="status",
            q=q,
        ),
    }


def get_budget_breakdown_for_department(
    department_name: str,
    year: int | None = None,
    group_by: str = "category",
    q: str | None = None,
) -> dict:
    rows = list_budget_records_for_department(department_name, year=year, q=q)

    buckets = defaultdict(
        lambda: {
            "label": "",
            "total_spent": Decimal("0"),
            "record_count": 0,
            "records": [],
        }
    )

    for row in rows:
        if group_by == "vendor":
            label = row.get("vendor_name") or "Unassigned Vendor"

        elif group_by == "month":
            anchor = row.get("_budget_anchor_date")
            label = anchor.strftime("%Y-%m") if anchor else "No Date"

        elif group_by == "record_type":
            label = (
                row.get("record_type") or "Unassigned Type"
            ).replace("_", " ").title()

        elif group_by == "status":
            label = (
                row.get("status") or "Unassigned Status"
            ).replace("_", " ").title()

        else:
            label = row.get("category_name") or "Unassigned Category"

        bucket = buckets[label]
        bucket["label"] = label
        bucket["total_spent"] += row["_budget_cost"]
        bucket["record_count"] += 1
        bucket["records"].append(row)

    total_spent = sum(
        (item["total_spent"] for item in buckets.values()),
        Decimal("0"),
    )

    results = []

    if group_by == "month":
        sorted_bucket_items = sorted(
            buckets.items(),
            key=lambda kv: kv[0],
        )
    else:
        sorted_bucket_items = sorted(
            buckets.items(),
            key=lambda kv: (kv[1]["total_spent"], kv[0]),
            reverse=True,
        )

    for label, item in sorted_bucket_items:
        average_spend = (
            item["total_spent"] / item["record_count"]
            if item["record_count"]
            else Decimal("0")
        )

        percent_of_total = (
            (item["total_spent"] / total_spent) * Decimal("100")
            if total_spent
            else Decimal("0")
        )

        results.append(
            {
                "label": item["label"],
                "total_spent": item["total_spent"],
                "record_count": item["record_count"],
                "average_spend": average_spend,
                "percent_of_total": percent_of_total,
            }
        )

    chart_labels = [item["label"] for item in results]
    chart_values = [float(item["total_spent"]) for item in results]

    top_bucket = results[0] if results else None

    return {
        "rows": results,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "top_bucket": top_bucket,
        "group_by": group_by,
    }

def get_budget_target_for_department(
    department_name: str,
    fiscal_year: int | None,
) -> dict:
    department_name = normalize_text(department_name)

    if not department_name or not fiscal_year:
        return {
            "department_name": department_name or "",
            "fiscal_year": fiscal_year,
            "total_budget": Decimal("0"),
            "notes": "",
        }

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM finance_budget_targets
            WHERE department_name = ?
              AND fiscal_year = ?
            """,
            (department_name, fiscal_year),
        ).fetchone()

    if not row:
        return {
            "department_name": department_name,
            "fiscal_year": fiscal_year,
            "total_budget": Decimal("0"),
            "notes": "",
        }

    item = dict(row)
    item["total_budget"] = parse_cost_to_decimal(item.get("total_budget"))
    return item


def save_budget_target_for_department(
    *,
    department_name: str,
    fiscal_year: int,
    total_budget: str | None,
    notes: str | None = None,
    created_by_user_id: int | None = None,
):
    department_name = normalize_text(department_name)
    if not department_name:
        raise ValueError("department_name is required")

    if not fiscal_year:
        raise ValueError("fiscal_year is required")

    total_budget_decimal = parse_cost_to_decimal(total_budget)
    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO finance_budget_targets (
                department_name,
                fiscal_year,
                total_budget,
                notes,
                created_by_user_id,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(department_name, fiscal_year) DO UPDATE SET
                total_budget = excluded.total_budget,
                notes = excluded.notes,
                updated_at = excluded.updated_at
            """,
            (
                department_name,
                fiscal_year,
                str(total_budget_decimal.quantize(Decimal("0.01"))),
                normalize_text(notes),
                created_by_user_id,
                now,
                now,
            ),
        )
        conn.commit()

def validate_transactions_import(
    *,
    run_id: int,
    profile_id: int,
    default_department_name: str | None = None,
    preview_limit: int = 20,
) -> dict:
    run = get_import_run_by_id(run_id)
    if not run:
        raise ValueError("Import run not found")

    rows = read_import_rows(run["stored_filename"])
    mappings = get_import_profile_fields(profile_id)

    clear_import_run_errors(run_id)

    preview_rows = []
    errors = []
    valid_rows = 0
    skipped_rows = 0
    vendors_to_create = 0

    for index, row in enumerate(rows, start=2):
        mapped = apply_import_mapping(row, mappings, index)

        if default_department_name:
            mapped["department_name"] = normalize_text(default_department_name)

        if should_skip_efinance_transaction(mapped, row):
            skipped_rows += 1
            continue

        mapped["transaction_type"] = classify_transaction_type(mapped)
        mapped["suggested_record_type"] = suggest_record_type_for_transaction(mapped)
        mapped["title"] = build_transaction_title(mapped, row, index)
        mapped["is_promotable"] = 1 if mapped.get("suggested_record_type") else 0

        vendor_name = normalize_text(mapped.get("vendor_name"))
        vendor_code = normalize_text(mapped.get("vendor_code"))

        if vendor_name or vendor_code:
            existing_vendor = find_vendor_for_import(
                vendor_name=vendor_name,
                vendor_code=vendor_code,
            )
            if not existing_vendor and vendor_name:
                vendors_to_create += 1

        valid_rows += 1

        if len(preview_rows) < preview_limit:
            preview_rows.append(mapped)

    return {
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "skipped_rows": skipped_rows,
        "vendors_to_create": vendors_to_create,
        "preview_rows": preview_rows,
        "errors": errors,
    }


def execute_transactions_import(
    *,
    run_id: int,
    profile_id: int,
    default_department_name: str | None = None,
    created_by_user_id: int | None = None,
) -> dict:
    run = get_import_run_by_id(run_id)
    if not run:
        raise ValueError("Import run not found")

    rows = read_import_rows(run["stored_filename"])
    mappings = get_import_profile_fields(profile_id)

    total_rows = len(rows)
    created_rows = 0
    skipped_rows = 0
    error_rows = 0
    vendors_created = 0

    update_import_run_results(
        run_id,
        status="running",
        total_rows=total_rows,
        created_rows=0,
        updated_rows=0,
        skipped_rows=0,
        error_rows=0,
        run_notes="Transaction import execution started.",
        completed=False,
    )

    for index, row in enumerate(rows, start=2):
        try:
            mapped = apply_import_mapping(row, mappings, index)

            if default_department_name:
                mapped["department_name"] = normalize_text(default_department_name)

            if should_skip_efinance_transaction(mapped, row):
                skipped_rows += 1
                continue

            transaction_type = classify_transaction_type(mapped)
            suggested_record_type = suggest_record_type_for_transaction(mapped)
            title = build_transaction_title(mapped, row, index)

            vendor_id, vendor_was_created = get_or_create_vendor_for_import(
                vendor_name=mapped.get("vendor_name"),
                vendor_code=mapped.get("vendor_code"),
            )

            if vendor_was_created:
                vendors_created += 1

            create_finance_transaction(
                department_name=mapped.get("department_name") or "",
                import_run_id=run_id,
                source_type=run.get("source_type") or "manual_upload",
                source_row_number=index,
                transaction_type=transaction_type,
                review_status="needs_review",
                title=title,
                description=mapped.get("description"),
                vendor_id=vendor_id,
                vendor_code=mapped.get("vendor_code"),
                vendor_name=mapped.get("vendor_name"),
                fund=mapped.get("fund"),
                budget_unit=mapped.get("budget_unit"),
                account_code=mapped.get("account_code"),
                po_number=mapped.get("po_number"),
                purchase_date=mapped.get("purchase_date"),
                expenditure_amount=mapped.get("expenditure_amount"),
                encumbrance_amount=mapped.get("encumbrance_amount"),
                cumulative_balance=mapped.get("cumulative_balance"),
                suggested_record_type=suggested_record_type,
                is_promotable=bool(suggested_record_type),
                raw_json=json.dumps(row),
            )

            created_rows += 1

        except Exception as exc:
            error_rows += 1
            log_import_run_error(
                run_id=run_id,
                row_number=index,
                source_identifier=row.get(next(iter(row.keys()), ""), "") if row else None,
                error_message=str(exc),
            )

    run_notes = (
        f"Transaction import completed. Created transactions: {created_rows}, "
        f"Created vendors: {vendors_created}, Skipped: {skipped_rows}, Errors: {error_rows}."
    )

    final_status = "completed"
    if error_rows and not created_rows:
        final_status = "completed_with_errors"
    elif error_rows:
        final_status = "completed_with_warnings"

    update_import_run_results(
        run_id,
        status=final_status,
        total_rows=total_rows,
        created_rows=created_rows,
        updated_rows=0,
        skipped_rows=skipped_rows,
        error_rows=error_rows,
        run_notes=run_notes,
        completed=True,
    )

    return {
        "total_rows": total_rows,
        "created_rows": created_rows,
        "updated_rows": 0,
        "skipped_rows": skipped_rows,
        "error_rows": error_rows,
        "vendors_created": vendors_created,
        "notifications_sent": 0,
        "status": final_status,
        "run_notes": run_notes,
    }

def bulk_update_transactions_review_status(
    *,
    transaction_ids: list[int],
    department_name: str,
    review_status: str,
) -> int:
    allowed_statuses = {"needs_review", "ignored"}
    review_status = normalize_text(review_status)

    if review_status not in allowed_statuses:
        raise ValueError("Invalid review status.")

    clean_ids = [int(item) for item in transaction_ids if str(item).isdigit()]
    if not clean_ids:
        return 0

    now = utc_now_iso()
    placeholders = ",".join("?" for _ in clean_ids)

    with get_connection() as conn:
        cursor = conn.execute(
            f"""
            UPDATE finance_transactions
            SET review_status = ?,
                updated_at = ?
            WHERE department_name = ?
              AND review_status != 'promoted'
              AND id IN ({placeholders})
            """,
            [review_status, now, department_name, *clean_ids],
        )
        conn.commit()
        return cursor.rowcount
    
def normalize_category_lookup(value: str | None) -> str | None:
    value = normalize_text(value)
    if not value:
        return None

    value = value.lower()
    value = value.replace("&", " and ")
    value = value.replace("-", " ")
    value = value.replace("_", " ")
    value = value.replace("/", " ")

    cleaned = []
    for char in value:
        if char.isalnum() or char.isspace():
            cleaned.append(char)

    words = "".join(cleaned).split()
    normalized_words = []

    for word in words:
        if len(word) > 3 and word.endswith("ies"):
            word = word[:-3] + "y"
        elif len(word) > 3 and word.endswith("s"):
            word = word[:-1]

        normalized_words.append(word)

    return " ".join(normalized_words)


def ensure_category_alias(category_id: int, alias_name: str):
    alias_name = normalize_text(alias_name)
    normalized_alias = normalize_category_lookup(alias_name)

    if not category_id or not alias_name or not normalized_alias:
        return

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO finance_category_aliases (
                category_id,
                alias_name,
                normalized_alias,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (category_id, alias_name, normalized_alias, now, now),
        )
        conn.commit()


def find_category_suggestion_for_import(category_name: str | None) -> dict | None:
    category_name = normalize_text(category_name)
    normalized = normalize_category_lookup(category_name)

    if not category_name or not normalized:
        return None

    with get_connection() as conn:
        categories = conn.execute(
            """
            SELECT id, category_name
            FROM finance_categories
            WHERE is_active = 1
            """
        ).fetchall()

    best_category = None
    best_score = 0.0

    for category in categories:
        category_normalized = normalize_category_lookup(category["category_name"])
        if not category_normalized:
            continue

        score = SequenceMatcher(None, normalized, category_normalized).ratio()

        imported_words = set(normalized.split())
        category_words = set(category_normalized.split())

        if imported_words and category_words:
            overlap_score = len(imported_words & category_words) / len(
                imported_words | category_words
            )
            score = max(score, overlap_score)

        if score > best_score:
            best_score = score
            best_category = dict(category)

    if not best_category:
        return None

    now = utc_now_iso()
    confidence_score = round(best_score, 4)

    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO finance_category_import_suggestions (
                imported_category_name,
                normalized_imported_name,
                suggested_category_id,
                confidence_score,
                status,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, 'pending', ?, ?)
            ON CONFLICT(normalized_imported_name) DO UPDATE SET
                imported_category_name = excluded.imported_category_name,
                suggested_category_id = excluded.suggested_category_id,
                confidence_score = excluded.confidence_score,
                status = CASE
                    WHEN finance_category_import_suggestions.status = 'approved'
                    THEN finance_category_import_suggestions.status
                    ELSE 'pending'
                END,
                updated_at = excluded.updated_at
            """,
            (
                category_name,
                normalized,
                best_category["id"],
                confidence_score,
                now,
                now,
            ),
        )

        row = conn.execute(
            """
            SELECT
                s.*,
                c.category_name AS suggested_category_name
            FROM finance_category_import_suggestions s
            LEFT JOIN finance_categories c ON c.id = s.suggested_category_id
            WHERE s.normalized_imported_name = ?
            """,
            (normalized,),
        ).fetchone()

        conn.commit()

    return dict(row) if row else None

def get_category_import_suggestion_by_id(suggestion_id: int) -> dict | None:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT
                s.*,
                c.category_name AS suggested_category_name
            FROM finance_category_import_suggestions s
            LEFT JOIN finance_categories c ON c.id = s.suggested_category_id
            WHERE s.id = ?
            """,
            (suggestion_id,),
        ).fetchone()

    return dict(row) if row else None


def approve_category_import_suggestion(suggestion_id: int):
    suggestion = get_category_import_suggestion_by_id(suggestion_id)

    if not suggestion:
        raise ValueError("Category suggestion not found.")

    category_id = suggestion.get("suggested_category_id")
    imported_category_name = suggestion.get("imported_category_name")

    if not category_id:
        raise ValueError("Suggestion does not have a category match.")

    ensure_category_alias(
        category_id=category_id,
        alias_name=imported_category_name,
    )

    now = utc_now_iso()

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_category_import_suggestions
            SET status = 'approved',
                updated_at = ?
            WHERE id = ?
            """,
            (now, suggestion_id),
        )
        conn.commit()